import tkinter as tk
from tkinter import ttk, messagebox, filedialog, Listbox, SINGLE, END
import threading
import queue
import json
import uuid
import os
import datetime
import glob
import pandas as pd
from openpyxl import Workbook, load_workbook
import math  # Added for math.ceil in FVE calculation
from openpyxl.styles import PatternFill, Font, Alignment
import requests
import re
import sys
import subprocess
from pathlib import Path
from collections import Counter
from geopy.geocoders import Nominatim
from geopy.extra.rate_limiter import RateLimiter
import unicodedata
import shutil  # Added based on user's snippet, assuming it was intended as a separate import
import io  # Pre prácu s in-memory súbormi
try:
    from PIL import Image
except ImportError:
    messagebox.showerror(
        "Chýbajúca knižnica",
        "Knižnica 'Pillow' nie je nainštalovaná. Prosím, nainštalujte ju príkazom: pip install Pillow")
    sys.exit(1)
try:
    import customtkinter as ctk
except ImportError:
    messagebox.showerror(
        "Chýbajúca knižnica",
        "Knižnica 'customtkinter' nie je nainštalovaná. Prosím, nainštalujte ju príkazom: pip install customtkinter")
    sys.exit(1)
try:
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
except ImportError:
    messagebox.showerror(
        "Chýbajúca knižnica",
        "Knižnica 'matplotlib' nie je nainštalovaná. Prosím, nainštalujte ju príkazom: pip install matplotlib")
    sys.exit(1)
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.utils import ImageReader
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import Paragraph
except ImportError:
    messagebox.showerror(
        "Chýbajúca knižnica",
        "Knižnica 'reportlab' nie je nainštalovaná. Prosím, nainštalujte ju príkazom: pip install reportlab")
    sys.exit(1)

# =====================================================================================
#  ČASŤ 1: LOGIKA Z HLAVNÉHO ANALYTICKÉHO SKRIPTU
# =====================================================================================

from config import *  # Import all constants from config.py
import database_manager  # Import databázového manažéra

# --- KONŠTANTY PRE FARBY ZVÝRAZŇOVANIA (pre Excel) ---
ZLT_VYPLN = PatternFill(
    start_color="FFFF00",
    end_color="FFFF00",
    fill_type="solid")
ZELENA_VYPLN = PatternFill(
    start_color="90EE90",
    end_color="90EE90",
    fill_type="solid")
ORANZOVA_VYPLN = PatternFill(
    start_color="FFA500",
    end_color="FFA500",
    fill_type="solid")
MODRA_VYPLN = PatternFill(
    start_color="ADD8E6",
    end_color="ADD8E6",
    fill_type="solid")
FIALOVA_VYPLN = PatternFill(
    start_color="E0B0FF",
    end_color="E0B0FF",
    fill_type="solid")  # Nová farba pre záložný blok

if getattr(sys, 'frozen', False):
    ADRESAR_SKRIPTU = Path(os.path.dirname(sys.executable))
else:
    try:
        ADRESAR_SKRIPTU = Path(os.path.dirname(os.path.abspath(__file__)))
    except NameError:
        ADRESAR_SKRIPTU = Path.cwd()

VYSTUPNY_SUBOR_PREFIX = "vysledky_analyzy"

geolocator = Nominatim(user_agent="skript_analyzy_fve_gui")
geocode = RateLimiter(
    geolocator.geocode,
    min_delay_seconds=1,
    error_wait_seconds=5)


def ziskaj_okte_data(datum_na_stiahnutie, log_func):
    """
    Stiahne denné dáta z OKTE pre konkrétny dátum a vráti ich ako DataFrame.
    Neukladá súbor na disk.
    """
    date_string_api = datum_na_stiahnutie.strftime("%Y-%m-%d")
    date_display = datum_na_stiahnutie.strftime("%d.%m.%Y")

    log_func(f"    -> Pokúšam sa stiahnuť dáta pre: {date_display}")
    download_url = f"{BASE_DOWNLOAD_URL}?deliverydayfrom={date_string_api}&deliverydayto={date_string_api}&simple=false&lang=sk-SK&format=xlsx"

    for pokus in range(3):
        try:
            response = requests.get(download_url, stream=True, timeout=20)
            if response.status_code == 503:
                log_func(
                    f"      -> UPOZORNENIE: Server OKTE je dočasne nedostupný (503). Pokus {pokus + 1}/3.")
                if pokus < 2:
                    import time
                    time.sleep(5)
                    continue

            if response.status_code == 200 and response.content and len(
                    response.content) > 1000:
                try:
                    # Načítame Excel priamo z pamäte
                    df = pd.read_excel(io.BytesIO(response.content))
                    log_func(
                        f"      -> ✅ ÚSPECH: Dáta pre {date_display} boli stiahnuté do pamäte.")
                    return df
                except Exception as e:
                    log_func(
                        f"      -> ❌ CHYBA pri čítaní Excelu z pamäte: {e}")
                    return None
            elif response.status_code == 404:
                log_func(
                    f"      -> UPOZORNENIE: Dáta pre {date_display} nie sú dostupné (404).")
                return None
            else:
                log_func(
                    f"      -> UPOZORNENIE: Odpoveď servera je prázdna alebo nastala chyba (kód: {
                        response.status_code}).")
            break
        except requests.exceptions.RequestException as e:
            log_func(f"      -> ❌ CHYBA POŽIADAVKY pre {date_display}: {e}.")
            break
    return None


def obohat_data(df_okte, fve_15min_data, log_func):
    """
    Obohatí DataFrame o stĺpce 'cislo bloku' a 'predpoved FVE'.
    Pracuje priamo s DataFrame v pamäti.
    """
    try:
        # Kontrola stĺpcov
        if STLPEC_CAS not in df_okte.columns or STLPEC_CIEN not in df_okte.columns:
            log_func(
                f"       CHYBA: V stiahnutých dátach chýba stĺpec '{STLPEC_CAS}' alebo '{STLPEC_CIEN}'.")
            return None

        df = df_okte.copy()
        # Pridanie stĺpcov
        df['cislo bloku'] = 0
        df[STLPEC_FVE_KWH] = 0.0

        for idx in range(len(df)):
            # FVE predpoveď
            fve_val = fve_15min_data.get(idx, 0)
            df.at[idx, STLPEC_FVE_KWH] = fve_val

            # Číslo bloku (6-hodinové bloky, 1-6)
            block_num = math.ceil((idx + 1) / POCET_INTERVALOV_V_6H_BLOKU)
            df.at[idx, 'cislo bloku'] = block_num

        log_func(f"       ✅ Dáta boli obohatené o FVE predpoveď a bloky.")
        return df
    except Exception as e:
        log_func(f"       CHYBA pri obohacovaní dát: {e}")
        return None


def zisti_datum_objekt_zo_suboru(nazov_suboru):
    nazov_suboru = os.path.basename(nazov_suboru)
    match = re.search(r'(\d{2}-\d{2}-\d{4})', nazov_suboru)
    if match:
        try:
            return datetime.datetime.strptime(match.group(1), "%d-%m-%Y")
        except ValueError:
            pass
    return None


def skore_na_text(skore_cislo):
    if skore_cislo is None:
        return "N/A"
    if skore_cislo <= 20:
        return "veľmi slabé"
    if skore_cislo <= 40:
        return "slabé"
    if skore_cislo <= 60:
        return "priemerné"
    if skore_cislo <= 80:
        return "dobré"
    return "výborné"


def vytvor_tabulku_vysledkov(log_func):
    df_vysledky = pd.DataFrame(columns=STLPCE_TABULKY)
    log_func(f"    -> Pripravená nová tabuľka výsledkov v pamäti.")
    return df_vysledky


def uprav_sirku_stlpcov(subor_vystup, log_func):
    """Automaticky upraví šírku stĺpcov vo všetkých hárkoch Excel súboru podľa obsahu."""
    try:
        workbook = load_workbook(subor_vystup)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for col in worksheet.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                adjusted_width = max_length + 4  # Pridáme malú rezervu
                worksheet.column_dimensions[column_letter].width = adjusted_width
        workbook.save(subor_vystup)
        log_func(
            f"    -> Šírka stĺpcov bola automaticky upravená vo všetkých hárkoch.")
    except Exception as e:
        log_func(
            f"       UPOZORNENIE: Nepodarilo sa upraviť šírku stĺpcov: {e}")


def ziskaj_predpoved_fve(latitude, longitude, datum_analyzy_str, log_func):
    """
    Získa 15-minútovú predpoveď FVE (0-100%) a celkové textové skóre.
    Vráti slovník s 15-minútovou výrobou (0-100%) a celkové textové skóre.
    """
    for pokus in range(3):
        try:
            datum_api = datetime.datetime.strptime(
                datum_analyzy_str, "%d.%m.%Y").strftime("%Y-%m-%d")
            url = f"{OPEN_METEO_API_URL}?latitude={latitude}&longitude={longitude}&hourly=cloudcover,shortwave_radiation,temperature_2m,direct_radiation,diffuse_radiation&start_date={datum_api}&end_date={datum_api}&timezone=auto"
            r = requests.get(url, timeout=15)
            r.raise_for_status()
            data = r.json().get('hourly', {})
            if not data.get('time'):
                return {}, "N/A"

            # Open-Meteo poskytuje hodinové dáta, potrebujeme 15-minútové.
            # Zjednodušene interpolujeme alebo použijeme hodinovú hodnotu pre všetky 15-minútové intervaly v danej hodine.
            # Pre tento prípad použijeme hodinovú hodnotu pre všetky
            # 15-minútové intervaly v danej hodine.

            # Kľúč: index intervalu (0-95), Hodnota: % výkonu
            fve_15min_prediction = {}

            # Pre výpočet maximálneho možného výkonu (100%)
            # Predpokladáme, že 1000 W/m^2 shortwave_radiation je ideálna podmienka pre PEAK_POWER_KW
            # A že 100% výkonu je dosiahnuté pri ideálnych podmienkach.

            # Zjednodušený výpočet:
            # 1. Odhadneme "ideálny" maximálny shortwave_radiation pre danú lokalitu a čas (ťažké bez komplexného modelu).
            # 2. Alebo použijeme fixnú referenčnú hodnotu, napr. 1000 W/m^2 ako 100%.
            # Použijeme druhú možnosť pre jednoduchosť.
            MAX_SOLAR_RADIATION = 1000  # W/m^2, referenčná hodnota pre 100% výkon

            for i, t_str in enumerate(data['time']):
                hodina = datetime.datetime.fromisoformat(t_str).hour
                # shortwave_radiation je v W/m^2
                solar_radiation = data['shortwave_radiation'][i]
                cloud_cover = data['cloudcover'][i]  # 0-100%

                # Zjednodušený odhad FVE výkonu v % (0-100)
                # Zohľadníme slnečné žiarenie a oblačnosť
                # Výkon = (solar_radiation / MAX_SOLAR_RADIATION) * (1 - cloud_cover / 100) * 100
                # Alebo jednoduchšie: len na základe solar_radiation

                # Nová logika pre FVE skóre (0-100%)
                # Vychádzame z toho, že 1000 W/m^2 je 100% výkon.
                # A oblačnosť znižuje výkon.

                # Vypočítame potenciálny výkon na základe slnečného žiarenia
                potential_power_percent = (
                    solar_radiation / MAX_SOLAR_RADIATION) * 100

                # Znížime ho na základe oblačnosti (jednoduchý model)
                # Napr. 100% oblačnosť zníži výkon o 50% (toto je len príklad, dá sa to spresniť)
                # Alebo priamo použijeme cloudcover ako faktor zníženia

                # Použijeme jednoduchší model: FVE výkon je priamo úmerný shortwave_radiation
                # a mierne korigovaný oblačnosťou.
                # Max 100%
                # Delíme 200, aby oblačnosť nemala príliš veľký vplyv
                fve_percent = min(
                    100, max(0, round(potential_power_percent * (1 - cloud_cover / 200))))

                # Aplikujeme túto hodinovú hodnotu na všetky 15-minútové
                # intervaly v danej hodine
                for j in range(4):  # 4 x 15 minút v hodine
                    interval_index = hodina * 4 + j
                    fve_15min_prediction[interval_index] = fve_percent

            # Výpočet celkového textového skóre pre deň (priemer FVE výkonu)
            if fve_15min_prediction:
                avg_fve_percent = sum(
                    fve_15min_prediction.values()) / len(fve_15min_prediction)
                textove_skore = skore_na_text(round(avg_fve_percent))
            else:
                textove_skore = "N/A"

            return fve_15min_prediction, textove_skore

        except requests.exceptions.RequestException as e:
            log_func(
                f"       Chyba pri získavaní FVE predpovede (pokus {
                    pokus + 1}/3): {e}")
            if pokus < 2:
                log_func("       Čakám 5 sekúnd pred ďalším pokusom...")
                import time
                time.sleep(5)

    log_func(
        "       ❌ Nepodarilo sa získať FVE predpoveď ani po opakovaných pokusoch.")
    return {}, "N/A"


def ziskaj_pocasie_data(latitude, longitude, datum_analyzy, log_func):
    """
    Stiahne predpoveď počasia pre konkrétny deň a vráti ju ako DataFrame.
    Neukladá súbor na disk.
    """
    try:
        datum_obj = datetime.datetime.strptime(datum_analyzy, "%d.%m.%Y")
        datum_api = datum_obj.strftime("%Y-%m-%d")
        nazov_harku = datum_obj.strftime("%d.%m.%Y")
    except ValueError:
        log_func(
            f"   -> UPOZORNENIE: Neplatný formát dátumu '{datum_analyzy}' pre predpoveď počasia.")
        return None

    url = f"https://api.open-meteo.com/v1/forecast?latitude={latitude}&longitude={longitude}&hourly=cloudcover,shortwave_radiation,temperature_2m,direct_radiation,diffuse_radiation&start_date={datum_api}&end_date={datum_api}&timezone=auto"

    try:
        r = requests.get(url, timeout=15)
        r.raise_for_status()
        data = r.json().get('hourly', {})
        if not data.get('time'):
            log_func(
                f"   -> UPOZORNENIE: API pre počasie na deň {nazov_harku} nevrátilo žiadne dáta.")
            return None

        df_pocasie = pd.DataFrame(data)
        log_func(
            f"   -> ✅ Predpoveď počasia pre {nazov_harku} bola úspešne stiahnutá do pamäte.")
        return df_pocasie

    except requests.exceptions.RequestException as e:
        log_func(f"   -> ❌ CHYBA pri sťahovaní predpovede počasia: {e}")
        return None


def najdi_optimalny_blok_dynamicky(
        df,
        dlzka_v_intervaloch,
        typ_bloku,
        casovy_rozsah_indexov=None,
        vylucit_indexy=None):
    """
    Nájdeme optimálny blok s dynamickou dĺžkou v rámci celého dňa alebo zadaného rozsahu.
    """
    best_avg_price = float('inf') if typ_bloku == 'nakup' else -float('inf')
    best_block_indices = []

    df_search = df.copy()
    if casovy_rozsah_indexov:
        df_search = df_search.iloc[casovy_rozsah_indexov[0]:casovy_rozsah_indexov[1]]

    if vylucit_indexy:
        df_search = df_search.drop(index=vylucit_indexy, errors='ignore')

    if df_search.empty or len(df_search) < dlzka_v_intervaloch:
        return "N/A", 0.0, []

    # Vytvoríme kĺzavý priemer
    rolling_avg = df_search[STLPEC_CIEN].rolling(
        window=dlzka_v_intervaloch).mean()

    if typ_bloku == 'nakup':
        if rolling_avg.empty or rolling_avg.isna().all():
            return "N/A", 0.0, []
        best_end_index_local = rolling_avg.idxmin()
    else:  # predaj
        if rolling_avg.empty or rolling_avg.isna().all():
            return "N/A", 0.0, []
        best_end_index_local = rolling_avg.idxmax()

    # Prepočet lokálneho indexu na globálny index pôvodného DataFrame
    best_end_index_global = df.index.get_loc(best_end_index_local)
    start_index_global = best_end_index_global - dlzka_v_intervaloch + 1

    best_block_indices = list(
        range(
            start_index_global,
            best_end_index_global +
            1))
    best_block_df = df.loc[best_block_indices]

    start_time = best_block_df[STLPEC_CAS].iloc[0].split(' - ')[0]
    end_time = best_block_df[STLPEC_CAS].iloc[-1].split(' - ')[1]
    best_block_interval = f"{start_time} - {end_time}"

    best_avg_price = round(best_block_df[STLPEC_CIEN].mean(), 2)

    return best_block_interval, best_avg_price, best_block_indices


def najdi_nulove_a_zaporne_ceny(df, stlpec_cien, stlpec_cas):
    nulove_ceny = df[df[stlpec_cien] <= 0].copy()  # type: ignore
    if nulove_ceny.empty:
        return ""

    # Upravené funkcie pre extrakciu času z formátu "HH:MM - HH:MM"
    def get_start_minutes(cas_interval):
        try:
            match = re.search(
                r'(\d{2}):(\d{2})',
                str(cas_interval).split(' - ')[0])
            if match:
                return int(match.group(1)) * 60 + int(match.group(2))
        except BaseException:
            pass
        return None

    def get_end_minutes(cas_interval):
        try:
            match = re.search(
                r'(\d{2}):(\d{2})',
                str(cas_interval).split(' - ')[1])
            if match:
                h, m = map(int, match.groups())
                # Ak je koniec 00:00, ale začiatok bol predchádzajúci deň,
                # znamená to 24:00
                if h == 0 and m == 0 and '23:45' in str(cas_interval):
                    return 24 * 60
                return h * 60 + m
        except BaseException:
            pass
        return None

    nulove_ceny['Start_Minutes'] = nulove_ceny[stlpec_cas].apply(
        get_start_minutes)
    nulove_ceny['End_Minutes'] = nulove_ceny[stlpec_cas].apply(get_end_minutes)
    nulove_ceny.dropna(subset=['Start_Minutes', 'End_Minutes'], inplace=True)
    if nulove_ceny.empty:
        return ""

    # Zoskupenie súvislých blokov
    nulove_ceny['block_group'] = (nulove_ceny['Start_Minutes'].shift(
        1) != nulove_ceny['End_Minutes']).cumsum()

    vysledok: list[str] = []
    for _, group in nulove_ceny.groupby('block_group'):
        start_cas = group[stlpec_cas].iloc[0].split(' - ')[0]
        end_cas = group[stlpec_cas].iloc[-1].split(' - ')[1]
        vysledok.append(f"{start_cas} - {end_cas}")

    return " | ".join(vysledok) if vysledok else ""


def vycisti_excel_subor(subor_vstup):
    """Vyčistí staré formátovanie a súhrnné údaje z Excel súboru (bez logovania)."""
    try:
        wb = load_workbook(subor_vstup)
        ws = wb.active
        # Vyčistíme len formátovanie stĺpca cien, ak existuje
        stlpec_cien_cislo = None
        for col_idx, bunka in enumerate(ws[1], 1):
            if bunka.value == STLPEC_CIEN:
                stlpec_cien_cislo = col_idx
                break
        if stlpec_cien_cislo is not None:
            for row_idx in range(
                    2, ws.max_row + 1):  # Prejdeme všetky riadky s dátami
                ws.cell(
                    row=row_idx,
                    column=stlpec_cien_cislo).fill = PatternFill(
                    fill_type=None)

        # Odstránime staré súhrnné údaje pod tabuľkou
        # Predpokladáme, že súhrnné údaje začínajú od MAX_DATA_ROW_INDEX +
        # EXCEL_SUMMARY_START_ROW_OFFSET
        if ws.max_row >= MAX_DATA_ROW_INDEX + EXCEL_SUMMARY_START_ROW_OFFSET:
            ws.delete_rows(
                MAX_DATA_ROW_INDEX +
                EXCEL_SUMMARY_START_ROW_OFFSET,
                ws.max_row)

        wb.save(subor_vstup)
    except Exception:
        pass  # Ignorujeme chyby pri čistení, hlavná analýza je dôležitejšia


def extrahuj_interval_minuty(casovy_interval):
    """Pomocná funkcia na extrakciu minút z časového intervalu."""
    if pd.isna(casovy_interval) or not isinstance(casovy_interval, str):
        return None, None
    cas_matches = re.findall(r'(\d{2}):(\d{2})', casovy_interval)
    if len(cas_matches) == 2:
        try:
            h1, m1 = map(int, cas_matches[0])
            h2, m2 = map(int, cas_matches[1])
            start_min = h1 * 60 + m1
            end_min = h2 * 60 + m2
            if h2 == 0 and m2 == 0 and h1 != 0:
                end_min = 24 * 60  # Pre 24:00
            return start_min, end_min
        except BaseException:
            return None, None
    return None, None


def ziskaj_bloky_na_zvyráznenie(df, bloky_na_zvyráznenie):
    """
    Pripraví slovník s indexami a farbami pre zvýraznenie na základe nových blokov.
    bloky_na_zvyráznenie je slovník: {'typ_bloku': 'HH:MM - HH:MM', ...}
    """
    indexy_farieb = {}
    nulove_indexy = set(df[df[STLPEC_CIEN] <= 0].index)

    def parse_cas_pasmo(pasmo_str):
        if pasmo_str == "N/A" or not isinstance(pasmo_str, str):
            return -1, -1
        try:
            cas_matches = re.findall(r'(\d{2}):(\d{2})', pasmo_str)
            if len(cas_matches) < 2:
                return -1, -1
            h1, m1 = map(int, cas_matches[0])
            h2, m2 = map(int, cas_matches[-1])
            start_min = h1 * 60 + m1
            end_min = h2 * 60 + m2
            if h2 == 0 and m2 == 0 and h1 != 0:
                end_min = 24 * 60
            return start_min, end_min
        except BaseException:
            return -1, -1

    # Definícia farieb pre jednotlivé typy blokov
    farby_blokov = {
        "blok 1 (00-06) - Nákup": ZELENA_VYPLN,
        "blok 2 (06-12) - Predaj": MODRA_VYPLN,
        # Nová farba pre nabíjanie
        "blok 3 (12-18) - Nabíjanie": ORANZOVA_VYPLN,
        "blok 4 (18-24) - Predaj": FIALOVA_VYPLN,
    }

    parsed_bloky = {
        k: parse_cas_pasmo(v) for k,
        v in bloky_na_zvyráznenie.items() if k in farby_blokov}

    for index, row in df.iterrows():
        row_start_min, _ = extrahuj_interval_minuty(row[STLPEC_CAS])
        if row_start_min is None:
            continue

        # Zvýraznenie nulových/záporných cien
        if index in nulove_indexy:
            indexy_farieb[index] = ZLT_VYPLN

        # Zvýraznenie definovaných blokov
        for blok_key, (start_min, end_min) in parsed_bloky.items():
            if start_min != -1 and start_min <= row_start_min < end_min:
                indexy_farieb[index] = farby_blokov[blok_key]
                break  # Zvýrazníme len prvým nájdeným blokom

    return indexy_farieb


def aplikuj_zvýraznenie_do_excelu(subor_vstup, indexy_farieb, log_func):
    """Aplikuje farebné zvýraznenie do stiahnutého Excel súboru."""
    if not indexy_farieb:
        return
    try:
        wb = load_workbook(subor_vstup)
        ws = wb.active

        # Nájdenie stĺpcov pre cenu a FVE
        stlpec_cien_cislo = None
        stlpec_fve_cislo = None
        for col_idx, bunka in enumerate(ws[1], 1):
            if bunka.value == STLPEC_CIEN:
                stlpec_cien_cislo = col_idx
            elif bunka.value == STLPEC_FVE_KWH:
                stlpec_fve_cislo = col_idx

        if stlpec_cien_cislo is None:
            log_func(
                f"       UPOZORNENIE: Nenašiel sa stĺpec '{STLPEC_CIEN}' pre zvýraznenie.")
            return

        for df_index, farba in indexy_farieb.items():
            excel_row_idx = df_index + 2  # +2 lebo Excel je 1-indexovaný a má hlavičku
            if excel_row_idx <= MAX_DATA_ROW_INDEX + 1:  # Zvýrazňujeme len riadky s dátami
                # Podmienené zvýraznenie: Oranžová pre FVE, ostatné pre cenu
                if farba == ORANZOVA_VYPLN and stlpec_fve_cislo is not None:
                    ws.cell(
                        row=excel_row_idx,
                        column=stlpec_fve_cislo).fill = farba
                else:
                    ws.cell(
                        row=excel_row_idx,
                        column=stlpec_cien_cislo).fill = farba
        wb.save(subor_vstup)
    except Exception as e:
        log_func(f"       CHYBA pri zvýrazňovaní Excel súboru: {e}")


def zapis_sumarne_udaje(subor_vstup, vysledky_analyzy, log_func):
    """Zapíše súhrnné údaje pod hlavnú tabuľku v Excel súbore."""
    try:
        wb = load_workbook(subor_vstup)
        ws = wb.active
        bold_font = Font(bold=True)

        # Nájdeme riadok, kde začínajú súhrnné údaje
        start_row = MAX_DATA_ROW_INDEX + EXCEL_SUMMARY_START_ROW_OFFSET

        # Vyčistíme staré súhrnné údaje, ak nejaké sú
        if ws.max_row >= start_row:
            for r_idx in range(start_row, ws.max_row + 1):
                for c_idx in range(1, ws.max_column + 1):
                    ws.cell(row=r_idx, column=c_idx).value = None
                    ws.cell(
                        row=r_idx,
                        column=c_idx).font = Font()  # Reset fontu
                    # Reset zarovnania
                    ws.cell(row=r_idx, column=c_idx).alignment = Alignment()

        udaje = [
            ("Blok 1 (00-06) - Nákup:",
             vysledky_analyzy.get("blok 1 (00-06) - Nákup", "N/A")),
            ("Cena Nákup 1 (€):", vysledky_analyzy.get("cena nákup 1 (€)", 0.0)),
            ("Blok 2 (06-12) - Predaj:",
             vysledky_analyzy.get("blok 2 (06-12) - Predaj", "N/A")),
            ("Cena Predaj 1 (€):", vysledky_analyzy.get("cena predaj 1 (€)", 0.0)),
            ("Blok 3 (12-18) - Nabíjanie:",
             vysledky_analyzy.get("blok 3 (12-18) - Nabíjanie", "N/A")),
            ("Cena Nákup 2 (€):", vysledky_analyzy.get("cena nákup 2 (€)", 0.0)),
            ("Blok 4 (18-24) - Predaj:",
             vysledky_analyzy.get("blok 4 (18-24) - Predaj", "N/A")),
            ("Cena Predaj 2 (€):", vysledky_analyzy.get("cena predaj 2 (€)", 0.0)),
            (None, None),  # Prázdny riadok
            ("Celkový zisk (€):", vysledky_analyzy.get("celkový zisk (€)", 0.0)),
            ("Predpoveď FVE (text):", vysledky_analyzy.get(
                "predpoved vykonu FVE (text)", "N/A")),
            ("Časové pásmo ≤ 0 €:", vysledky_analyzy.get("časové pásmo ≤ 0 €", "N/A"))
        ]

        for i, (popis, hodnota) in enumerate(udaje):  # type: ignore
            if popis is None:
                continue
            ws.cell(row=start_row + i, column=1, value=popis).font = bold_font
            bunka_hodnota = ws.cell(row=start_row + i, column=4, value=hodnota)
            bunka_hodnota.alignment = Alignment(
                horizontal='center', vertical='center')
            if isinstance(hodnota, (int, float)):
                bunka_hodnota.number_format = '0.00'
        wb.save(subor_vstup)
    except Exception as e:
        log_func(f"       CHYBA pri zápise súhrnných údajov: {e}")


def normalizuj_mesto(mesto_input):
    """
    Normalizuje názov slovenského mesta/obce.
    Vstup: "zilina", "KOSICE", "banska bystrica"
    Výstup: "Žilina", "Košice", "Banská Bystrica"
    """
    if not mesto_input:
        return mesto_input

    # 1. Skúsime priamo z lowercase verzie (ak užívateľ zadal presne ale bez
    # diakritiky)
    mesto_lower = mesto_input.lower().strip()
    if mesto_lower in SLOVENSKE_MESTA:
        return SLOVENSKE_MESTA[mesto_lower]

    # 2. Odstránime diakritiku zo vstupu a skúsime nájsť v slovníku
    # (Slovník má kľúče bez diakritiky)
    mesto_bez_diakritiky = ''.join(
        c for c in unicodedata.normalize('NFD', mesto_lower)
        if unicodedata.category(c) != 'Mn'
    )

    if mesto_bez_diakritiky in SLOVENSKE_MESTA:
        return SLOVENSKE_MESTA[mesto_bez_diakritiky]

    # 3. Fallback: Ak nenájdeme, použijeme .title()
    return mesto_input.title()


def process_client_analysis(
        mesto,
        klient_adresar_nazov,
        meno,
        priezvisko,
        battery_capacity_kwh,
        log_func,
        temp_info_log,
        target_date=None):
    log_func("\n" + "=" * 70)
    log_func(
        f"ZAČÍNAM SPRACOVANIE PRE KLIENTA: {klient_adresar_nazov} (Mesto: {
            mesto.upper()})")
    log_func("=" * 70)

    # Vytvorenie adresára pre klienta (už len pre prípadné PDF, xlsx sa
    # nerobia)
    klient_root_dir = ADRESAR_SKRIPTU / \
        normalizuj_mesto(mesto) / klient_adresar_nazov
    klient_root_dir.mkdir(parents=True, exist_ok=True)

    if target_date is None:
        dnes = datetime.date.today()
        target_date = dnes + datetime.timedelta(days=1)

    datum_analyzy_str = target_date.strftime("%d.%m.%Y")

    # Normalizácia názvu mesta pre korektné geokódovanie (napr. sala -> Šaľa)
    mesto_normalizovane = normalizuj_mesto(mesto)

    log_func(
        f"\nKrok 1: Získavam súradnice pre lokalitu: '{mesto_normalizovane}'")
    location = geocode(mesto_normalizovane, country_codes="SK")
    if not location:
        log_func(
            f"❌ CHYBA: Lokalitu '{mesto_normalizovane}' sa nepodarilo nájsť.")
        return None
    latitude, longitude, lokalita_nazov = location.latitude, location.longitude, location.address

    # Stiahnutie OKTE dát pre zvolený dátum (do pamäte)
    log_func(f"\nKrok 2: Sťahujem dáta z OKTE pre {datum_analyzy_str}...")
    df_okte = ziskaj_okte_data(target_date, log_func)
    if df_okte is None:
        log_func(f"❌ CHYBA: Nepodarilo sa stiahnuť OKTE dáta.")
        return None

    # Stiahnutie predpovede počasia (do pamäte)
    # Poznámka: FVE predpoveď potrebuje len súradnice, ale ziskaj_pocasie_data sťahuje detailné počasie
    # ktoré momentálne nepoužívame pre výpočet FVE (používame pysolar v ziskaj_predpoved_fve),
    # ale môžeme ho chcieť uložiť do DB pre budúcnosť?
    # Pre zjednodušenie teraz preskočíme sťahovanie detailného počasia z OpenMeteo,
    # lebo ziskaj_predpoved_fve si to rieši sama (alebo používa pysolar).
    # Ale ziskaj_predpoved_fve (riadok 263) používa len pysolar a cloudcover z OpenMeteo?
    # Pozrime sa na ziskaj_predpoved_fve: tá volá API OpenMeteo.

    # Získanie FVE predpovede (0-100% pre 15-min intervaly)
    log_func("\nKrok 3: Získavam predpoveď výkonu FVE...")
    fve_15min_data, pocasie_skore_text = ziskaj_predpoved_fve(
        latitude, longitude, datum_analyzy_str, log_func)
    if not fve_15min_data:
        log_func(f"❌ CHYBA: Nepodarilo sa získať FVE predpoveď.")
        return None

    # Obohatenie dát
    log_func("\nKrok 4: Obohacujem dáta...")
    df_final = obohat_data(df_okte, fve_15min_data, log_func)
    if df_final is None:
        log_func(f"❌ CHYBA: Nepodarilo sa obohatiť dáta.")
        return None

    # Analytická časť
    log_func("\nKrok 5: Spúšťam analýzu blokov...")
    klient_id = klient_adresar_nazov.split('_')[0]
    vysledky_dna, hourly_data_json = analyzuj_denne_data(
        df_final, klient_id, meno, priezvisko, lokalita_nazov, datum_analyzy_str, pocasie_skore_text, battery_capacity_kwh, log_func)

    log_func("\nKrok 6: Ukladám výsledky do databázy...")
    if vysledky_dna:
        try:
            # Konverzia dátumu na ISO formát pre DB
            parts = datum_analyzy_str.split('.')
            iso_date = f"{parts[2]}-{parts[1]}-{parts[0]}"

            if database_manager.save_daily_result(
                    klient_id, iso_date, vysledky_dna, hourly_data_json):
                log_func(
                    f"    -> ✅ Výsledok pre {datum_analyzy_str} bol uložený do databázy.")
            else:
                log_func(
                    f"    -> ❌ CHYBA: Nepodarilo sa uložiť výsledok do databázy.")
        except Exception as e:
            log_func(f"    -> ❌ CHYBA pri ukladaní do DB: {e}")
    else:
        log_func("    -> Žiadne nové výsledky na zápis.")

    log_func(f"✅ SPRACOVANIE PRE KLIENTA '{klient_adresar_nazov}' DOKONČENÉ.")
    return True


def analyzuj_denne_data(
        df,
        klient_id,
        meno,
        priezvisko,
        lokalita_nazov,
        datum_analyzy,
        pocasie_skore_text,
        battery_capacity_kwh,
        log_func):
    """
    Spracuje DataFrame a nájde optimálne nákupné/predajné bloky.
    Vráti (vysledky_dna, hourly_data_json).
    """
    log_func(f"\nANALÝZA DÁT pre dátum {datum_analyzy}")

    try:
        # Čistenie dát
        df[STLPEC_CIEN] = pd.to_numeric(df[STLPEC_CIEN], errors='coerce')
        df[STLPEC_FVE_KWH] = pd.to_numeric(df[STLPEC_FVE_KWH], errors='coerce')
        df.dropna(subset=[STLPEC_CIEN, STLPEC_CAS], inplace=True)
        if df.empty:
            log_func(f"       UPOZORNENIE: Dáta sú prázdne.")
            return None, None
    except Exception as e:
        log_func(f"       CHYBA pri príprave dát: {e}")
        return None, None

    vysledky_dna = {
        "id_klienta": klient_id,
        "meno": meno,
        "priezvisko": priezvisko,
        "lokalita": lokalita_nazov.split(',')[0].strip(),
        "dátum": datum_analyzy,
        "blok 1 (00-06) - Nákup": "N/A",
        "cena nákup 1 (€)": 0.0,
        "blok 2 (06-12) - Predaj": "N/A",
        "cena predaj 1 (€)": 0.0,
        "blok 3 (12-18) - Nabíjanie": "N/A",
        "cena nákup 2 (€)": 0.0,
        "blok 4 (18-24) - Predaj": "N/A",
        "cena predaj 2 (€)": 0.0,
        "celkový zisk (€)": 0.0,
        "predpoved vykonu FVE (text)": pocasie_skore_text,
        "časové pásmo ≤ 0 €": najdi_nulove_a_zaporne_ceny(
            df.copy(),
            STLPEC_CIEN,
            STLPEC_CAS)}
    celkovy_zisk = 0.0
    obsadene_indexy = []
    akcie_mapa = {}  # idx -> typ akcie (nakup, predaj, nabijanie)

    # --- CYKLUS 1: Nočný nákup a ranný predaj ---
    log_func("    -> Hľadám optimálny nočný nákup a ranný predaj...")
    cas_nakup_1, cena_nakup_1, indexy_nakup_1 = najdi_optimalny_blok_dynamicky(
        df, POCET_INTERVALOV_V_3H_BLOKU, 'nakup', casovy_rozsah_indexov=(0, 48))

    if indexy_nakup_1:
        start_predaj_1_idx = indexy_nakup_1[-1] + 1
        cas_predaj_1, cena_predaj_1, indexy_predaj_1 = najdi_optimalny_blok_dynamicky(
            df, POCET_INTERVALOV_V_3H_BLOKU, 'predaj', casovy_rozsah_indexov=(
                start_predaj_1_idx, 48))

        if indexy_predaj_1 and cena_predaj_1 > cena_nakup_1:
            log_func(
                f"       -> Ziskový obchod nájdený: Nákup za {
                    cena_nakup_1:.2f}€, Predaj za {
                    cena_predaj_1:.2f}€.")
            obsadene_indexy.extend(indexy_nakup_1)
            obsadene_indexy.extend(indexy_predaj_1)
            for idx in indexy_nakup_1:
                akcie_mapa[idx] = "nakup_1"
            for idx in indexy_predaj_1:
                akcie_mapa[idx] = "predaj_1"

            vysledky_dna["blok 1 (00-06) - Nákup"] = cas_nakup_1
            vysledky_dna["cena nákup 1 (€)"] = cena_nakup_1
            vysledky_dna["blok 2 (06-12) - Predaj"] = cas_predaj_1
            vysledky_dna["cena predaj 1 (€)"] = cena_predaj_1
            celkovy_zisk -= cena_nakup_1 * battery_capacity_kwh / 1000
            celkovy_zisk += cena_predaj_1 * battery_capacity_kwh / 1000 * UCINNOST_BATERIE
        else:
            log_func("       -> Prvý obchodný cyklus nie je ziskový.")

    # --- CYKLUS 2: Nabíjanie z FVE a večerný predaj ---
    log_func("    -> Hľadám optimálne nabíjanie z FVE a večerný predaj...")
    df['fve_rolling_sum'] = df[STLPEC_FVE_KWH].rolling(
        window=POCET_INTERVALOV_V_6H_BLOKU).sum()
    fve_charge_end_idx = df['fve_rolling_sum'].idxmax()
    fve_charge_start_idx = fve_charge_end_idx - POCET_INTERVALOV_V_6H_BLOKU + 1

    indexy_fve_nabijanie = list(
        range(
            fve_charge_start_idx,
            fve_charge_end_idx + 1))
    df_fve_blok = df.loc[indexy_fve_nabijanie]

    cas_fve_nabijanie_start = df_fve_blok[STLPEC_CAS].iloc[0].split(' - ')[0]
    cas_fve_nabijanie_end = df_fve_blok[STLPEC_CAS].iloc[-1].split(' - ')[1]
    cas_fve_nabijanie = f"{cas_fve_nabijanie_start} - {cas_fve_nabijanie_end}"

    vyrobena_energia_fve_kwh = (
        df_fve_blok[STLPEC_FVE_KWH].sum() / 100) * battery_capacity_kwh * 0.25
    energia_pre_predaj_2_kwh = min(
        battery_capacity_kwh,
        vyrobena_energia_fve_kwh)

    vysledky_dna["blok 3 (12-18) - Nabíjanie"] = cas_fve_nabijanie
    vysledky_dna["cena nákup 2 (€)"] = 0.0
    obsadene_indexy.extend(indexy_fve_nabijanie)
    for idx in indexy_fve_nabijanie:
        akcie_mapa[idx] = "nabijanie_fve"

    log_func(
        f"       -> Nabíjanie z FVE: {cas_fve_nabijanie}, vyrobené {
            vyrobena_energia_fve_kwh:.2f} kWh")

    # Hľadáme predaj v druhej polovici dňa
    cas_predaj_2, cena_predaj_2, indexy_predaj_2 = najdi_optimalny_blok_dynamicky(
        df, POCET_INTERVALOV_V_3H_BLOKU, 'predaj', casovy_rozsah_indexov=(
            48, 96), vylucit_indexy=obsadene_indexy)
    if indexy_predaj_2:
        obsadene_indexy.extend(indexy_predaj_2)
        for idx in indexy_predaj_2:
            akcie_mapa[idx] = "predaj_2"
        vysledky_dna["blok 4 (18-24) - Predaj"] = cas_predaj_2
        vysledky_dna["cena predaj 2 (€)"] = cena_predaj_2
        celkovy_zisk += cena_predaj_2 * energia_pre_predaj_2_kwh / 1000 * UCINNOST_BATERIE
        log_func(
            f"       -> Predaj 2: {cas_predaj_2} za {cena_predaj_2:.2f} €/MWh")

    vysledky_dna["celkový zisk (€)"] = round(celkovy_zisk, 2)

    # Generovanie JSON s hodinovými dátami pre vizualizáciu
    hourly_data = []
    for idx, row in df.iterrows():
        item = {
            "cas": row[STLPEC_CAS],
            "cena": float(row[STLPEC_CIEN]),
            "fve": float(row[STLPEC_FVE_KWH]),
            "akcia": akcie_mapa.get(idx, "nic")
        }
        hourly_data.append(item)

    log_func(f"       ✅ Analýza dokončená. Celkový zisk: {celkovy_zisk:.2f} €")

    return vysledky_dna, json.dumps(hourly_data)


def copy_analyzer_script_if_needed(client_dir_path, log_func):
    """
    Skontroluje, či existuje skript pre detailnú analýzu a v prípade potreby
    ho nakopíruje alebo aktualizuje z hlavnej šablóny.
    """
    analyzer_script_name = "detailed_analyzer.py"
    analyzer_path = client_dir_path / analyzer_script_name
    template_path = ADRESAR_SKRIPTU / analyzer_script_name

    log_func(f"\nKrok 9: Kontrolujem skript pre detailnú analýzu...")

    if not template_path.exists():
        log_func(
            f"    -> ❌ KRITICKÁ CHYBA: Chýba šablóna '{analyzer_script_name}' v hlavnom adresári. Nemôžem pokračovať.")
        return

    # Skontrolujeme, či je potrebné aktualizovať skript
    try:
        if not analyzer_path.exists() or os.path.getmtime(
                template_path) > os.path.getmtime(analyzer_path):  # type: ignore
            import shutil
            log_func(
                f"    -> Detekovaná nová alebo chýbajúca verzia. Kopírujem '{analyzer_script_name}'...")
            shutil.copy(template_path, analyzer_path)
            log_func(f"      -> ✅ Šablóna bola úspešne skopírovaná.")
    except Exception as e:
        log_func(
            f"    -> ❌ CHYBA pri kopírovaní skriptu pre detailnú analýzu: {e}")
        return

    # Krok 2: Spustenie skriptu v samostatnom procese
    log_func(f"    -> Spúšťam detailnú analýzu pre klienta...")
    try:
        subprocess.run([sys.executable,
                        str(analyzer_path)],
                       cwd=client_dir_path,
                       check=True,
                       capture_output=True,
                       text=True,
                       encoding='utf-8')
        # log_func(f"      -> ✅ Detailná analýza dokončená. Súbor 'summary.json' bol vytvorený/aktualizovaný.")
    except subprocess.CalledProcessError as e:
        log_func(f"    -> ❌ CHYBA pri spustení detailnej analýzy: {e.stderr}")
    except Exception as e:
        log_func(
            f"    -> ❌ NEOČAKÁVANÁ CHYBA pri spustení detailnej analýzy: {e}")


def run_main_analysis(log_func, target_dates=None):
    log_func("\n" + "=" * 70)
    log_func("SPÚŠŤAM HLAVNÚ ANALÝZU CIEN ENERGIÍ (DB verzia)")
    log_func("=" * 70)

    if target_dates is None:
        dnes = datetime.date.today()
        target_dates = [
            dnes +
            datetime.timedelta(
                days=1)]  # Pôvodne len zajtra

    try:
        conn = database_manager.get_db_connection()
        clients = conn.execute(
            "SELECT * FROM clients WHERE status='active'").fetchall()
        conn.close()

        log_func(f"\nNájdených {len(clients)} aktívnych klientov v databáze.")

        for datum in target_dates:
            datum_str = datum.strftime("%d.%m.%Y")
            log_func(f"\n--- SPRACOVÁVAM DEŇ: {datum_str} ---")

            for klient in clients:
                mesto = klient['city']
                klient_id = klient['id']
                meno = klient['first_name']
                priezvisko = klient['last_name']
                battery_capacity_kwh = klient['battery_capacity_kwh']

                if not all([mesto, klient_id, meno, priezvisko]):
                    continue

                klient_adresar_nazov = f"{klient_id}_{priezvisko}_{meno}"
                process_client_analysis(
                    mesto,
                    klient_adresar_nazov,
                    meno,
                    priezvisko,
                    battery_capacity_kwh,
                    log_func,
                    [],
                    target_date=datum)

        log_func(
            "\n" +
            "=" *
            25 +
            " Všetci klienti boli spracovaní. " +
            "=" *
            25)
        log_func("✅ ANALÝZA JE HOTOVÁ.")

    except Exception as e:
        log_func(f"\n❌ NEOČAKÁVANÁ GLOBÁLNA CHYBA pri spracovaní: {e}")

# =====================================================================================
#  ČASŤ 2: LOGIKA Z INŠTALAČNÉHO SKRIPTU (instalacny_skript.py)
# =====================================================================================

# --- Konštanty pre názvy hárkov ---
# AKTIVNI_SHEET a NEAKTIVNI_SHEET sú už importované z config.py
# VYSTUPNY_SUBOR_PREFIX je už importovaný z config.py


def run_add_client(
        city,
        first_name,
        last_name,
        battery_capacity_kwh,
        log_func):
    log_func("\n🎯 PRIDÁVAM NOVÉHO KLIENTA")
    log_func("=" * 55)

    if not all([city, first_name, last_name, battery_capacity_kwh]):
        log_func("❌ Musíte vyplniť všetky polia (mesto, meno, priezvisko, kapacita)!")
        return

    # Normalizácia mesta
    city = normalizuj_mesto(city)

    log_func("\nHľadám unikátne ID pre nového klienta...")

    # Generovanie ID (zjednodušené, kontrola cez DB by bola lepšia, ale uuid
    # stačí)
    while True:
        client_id = str(uuid.uuid4())[:4].upper()
        # Tu by sme mohli skontrolovať duplicitu v DB, ale pravdepodobnosť je
        # malá
        break

    if database_manager.add_client(
            client_id,
            first_name,
            last_name,
            city,
            battery_capacity_kwh):
        log_func(
            f"\n🎉 Klient '{first_name} {last_name}' bol úspešne vytvorený.")
        log_func("=" * 55)
        log_func("\n✅ Záznam bol automaticky pridaný do databázy.")
        return True
    else:
        log_func("\n❌ CHYBA: Nepodarilo sa pridať klienta do databázy.")
        return False


def run_delete_client(client_identifier, log_func):
    """Presunie klienta do stavu 'inactive' v databáze."""
    log_func(
        f"\n⚙️ Spúšťam proces odstránenia pre klienta: {client_identifier}")

    client_id = client_identifier.split('_')[0]
    conn = database_manager.get_db_connection()
    try:
        conn.execute(
            "UPDATE clients SET status='inactive' WHERE id=?", (client_id,))
        conn.commit()
        log_func(
            f"✅ ÚSPECH: Klient '{client_identifier}' bol presunutý do neaktívnych.")
    except Exception as e:
        log_func(f"❌ CHYBA pri deaktivácii klienta: {e}")
    finally:
        conn.close()


def run_restore_client(client_identifier, log_func):
    """Presunie klienta do stavu 'active' v databáze."""
    log_func(f"\n⚙️ Spúšťam proces obnovenia pre klienta: {client_identifier}")

    client_id = client_identifier.split('_')[0]
    conn = database_manager.get_db_connection()
    try:
        conn.execute(
            "UPDATE clients SET status='active' WHERE id=?", (client_id,))
        conn.commit()
        log_func(f"✅ ÚSPECH: Klient '{client_identifier}' bol obnovený.")
    except Exception as e:
        log_func(f"❌ CHYBA pri obnove klienta: {e}")
    finally:
        conn.close()


def run_permanently_delete_single_client(client_identifier, log_func):
    """Navždy vymaže konkrétneho klienta z databázy."""
    log_func(
        f"\n⚠️ Spúšťam TRVALÉ ODSTRÁNENIE klienta: {client_identifier}...")

    if database_manager.delete_client_permanently(client_identifier):
        log_func(
            f"✅ ÚSPECH: Klient '{client_identifier}' bol navždy vymazaný z databázy.")
    else:
        log_func(
            f"❌ CHYBA pri trvalom odstraňovaní klienta '{client_identifier}'.")


def run_edit_client(client_identifier, new_data, log_func):
    """Aktualizuje údaje existujúceho klienta v databáze."""
    log_func(f"\n⚙️ Spúšťam proces úpravy pre klienta: {client_identifier}")

    client_id = client_identifier.split('_')[0]
    conn = database_manager.get_db_connection()
    try:
        conn.execute('''
            UPDATE clients
            SET first_name=?, last_name=?, city=?, battery_capacity_kwh=?
            WHERE id=?
        ''', (
            new_data['meno'],
            new_data['priezvisko'],
            new_data['mesto'],
            new_data['kapacita batérie v kWh'],
            client_id
        ))
        conn.commit()
        log_func(
            f"✅ ÚSPECH: Údaje pre klienta '{client_identifier}' boli aktualizované.")
    except Exception as e:
        log_func(f"❌ CHYBA pri úprave klienta: {e}")
    finally:
        conn.close()


def get_client_list():
    """Načíta zoznam aktívnych klientov z databázy pre zobrazenie v GUI."""
    try:
        conn = database_manager.get_db_connection()
        clients = conn.execute(
            "SELECT * FROM clients WHERE status='active' ORDER BY last_name, first_name").fetchall()
        conn.close()
        return [f"{c['id']}_{c['last_name']}_{c['first_name']}" for c in clients]
    except Exception as e:
        print(f"Chyba pri načítaní klientov: {e}")
        return []


def get_inactive_client_list():
    """Načíta zoznam neaktívnych klientov z databázy pre zobrazenie v GUI."""
    try:
        conn = database_manager.get_db_connection()
        clients = conn.execute(
            "SELECT * FROM clients WHERE status='inactive' ORDER BY last_name, first_name").fetchall()
        conn.close()
        return [f"{c['id']}_{c['last_name']}_{c['first_name']}" for c in clients]
    except Exception as e:
        print(f"Chyba pri načítaní neaktívnych klientov: {e}")
        return []


def get_client_details(client_identifier):
    """Načíta detailné údaje jedného klienta z databázy."""
    try:
        client_id = client_identifier.split('_')[0]
        conn = database_manager.get_db_connection()
        client = conn.execute(
            "SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
        conn.close()

        if client:
            return {
                'id': client['id'],
                'meno': client['first_name'],
                'priezvisko': client['last_name'],
                'mesto': client['city'],
                'kapacita batérie v kWh': client['battery_capacity_kwh']
            }
    except Exception:
        return None
    return None


def open_directory(path, log_func):
    """Otvori daný adresár v prieskumníkovi súborov."""
    try:
        if not os.path.exists(path):
            log_func(f"❌ CHYBA: Adresár '{path}' neexistuje.")
            messagebox.showerror("Chyba", f"Adresár neexistuje:\n{path}")
            return

        # os.startfile je najjednoduchší spôsob na Windows
        if sys.platform == "win32":
            os.startfile(path)
        # Alternatívy pre iné systémy
        elif sys.platform == "darwin":  # macOS
            subprocess.Popen(["open", path])
        else:  # Linux
            subprocess.Popen(["xdg-open", path])

        log_func(f"✅ Adresár '{path}' bol otvorený.")

    except Exception as e:
        log_func(f"❌ CHYBA pri otváraní adresára: {e}")
        messagebox.showerror("Chyba", f"Nepodarilo sa otvoriť adresár:\n{e}")


def calculate_client_summary(client_dir_path_unused, log_func):
    """
    Načíta všetky výsledky pre klienta z databázy a vypočíta súhrnné štatistiky.
    Argument client_dir_path_unused je ignorovaný, ponechaný pre kompatibilitu.
    Miesto toho extrahujeme ID klienta z kontextu (hack: musíme ho dostať inak alebo ho extrahovať z path ak sa dá).

    POZOR: Táto funkcia bola pôvodne volaná s client_dir_path.
    Teraz ju voláme s cestou: ADRESAR_SKRIPTU / normalizuj_mesto(mesto) / client_identifier
    Takže client_identifier je posledná časť cesty.
    """
    try:
        client_identifier = os.path.basename(client_dir_path_unused)
        client_id = client_identifier.split('_')[0]

        log_func(
            f"Načítavam výsledky z databázy pre klienta ID: {client_id}...")
        results = database_manager.get_all_results_for_client(client_id)

        if not results:
            return (
                {"error": "V databáze pre tohto klienta zatiaľ nie sú žiadne výsledky."}, None)

        # Prevod na DataFrame pre ľahšiu manipuláciu
        df = pd.DataFrame(results)

        # Premenovanie stĺpcov pre kompatibilitu so zvyškom kódu (generovanie PDF, grafov)
        # DB stĺpce: total_profit, price_sell_1, price_sell_2, date, fve_prediction_text
        # Očakávané v GUI/PDF: 'celkový zisk (€)', 'cena predaj 1 (€)', 'cena
        # predaj 2 (€)', 'dátum', 'predpoved vykonu FVE (text)'

        rename_map = {
            'total_profit': 'celkový zisk (€)',
            'price_sell_1': 'cena predaj 1 (€)',
            'price_sell_2': 'cena predaj 2 (€)',
            'date': 'dátum',
            'fve_prediction_text': 'predpoved vykonu FVE (text)'
        }
        df.rename(columns=rename_map, inplace=True)

        profit_col = 'celkový zisk (€)'
        sales_price_cols = ["cena predaj 1 (€)", "cena predaj 2 (€)"]

        # Konverzia na čísla (pre istotu, hoci z DB by mali prísť ako
        # čísla/None)
        df[profit_col] = pd.to_numeric(
            df[profit_col], errors='coerce').fillna(0.0)
        for col in sales_price_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)

        total_profit = df[profit_col].sum()
        avg_profit = df[profit_col].mean()
        total_sales_prices = df[sales_price_cols[0]
                                ].sum() + df[sales_price_cols[1]].sum()
        profitable_days = (df[profit_col] > 0).sum()
        loss_days = (df[profit_col] < 0).sum()

        # Formátovanie dátumu pre zobrazenie (z YYYY-MM-DD na DD.MM.YYYY)
        # Ale pozor, df['dátum'] môže byť string YYYY-MM-DD.
        # Pre PDF potrebujeme DD.MM.YYYY.
        try:
            df['dátum'] = pd.to_datetime(df['dátum']).dt.strftime('%d.%m.%Y')
        except BaseException:
            pass

        best_day_idx = df[profit_col].idxmax()
        worst_day_idx = df[profit_col].idxmin()

        best_day_row = df.loc[best_day_idx]
        worst_day_row = df.loc[worst_day_idx]

        fve_stats = Counter(df['predpoved vykonu FVE (text)'].dropna())

        summary = {
            "Rozdiel nakup / predaj": f"{total_profit:.2f} €",
            "Priemerný denný zisk": f"{avg_profit:.2f} €",
            "Potenciálny brutto zisk (sum cien predaja)": f"{total_sales_prices:.2f} €",
            "Počet ziskových dní": str(profitable_days),
            "Počet stratových dní": str(loss_days),
            "Najlepší deň": f"{best_day_row['dátum']} ({best_day_row[profit_col]:.2f} €)",
            "Najhorší deň": f"{worst_day_row['dátum']} ({worst_day_row[profit_col]:.2f} €)",
            "Štatistika FVE": dict(fve_stats)
        }
        return summary, df

    except Exception as e:
        log_func(f"❌ CHYBA pri výpočte súhrnu z DB: {e}")
        return ({"error": f"Nastala chyba: {e}"}, None)

# =====================================================================================
#  ČASŤ 3: GRAFICKÉ ROZHRANIE (GUI)
# =====================================================================================


# --- FAREBNÁ PALETA A ŠTÝLY PODĽA LOGA ---
BG_COLOR = "#000000"          # Čierna z pozadia loga
LOG_BG_COLOR = "#1E1E1E"      # Veľmi tmavá šedá pre log
TEXT_COLOR = "#FFFFFF"        # Biela pre text
BUTTON_COLOR = "#FFC107"      # Mierne sýtejšia žltá
BUTTON_TEXT_COLOR = "#000000"  # Čierna pre text na tlačidlách
BUTTON_ACTIVE_BG = "#FFA000"  # Tmavšia žltá pri kliknutí

GREEN_COLOR = "#4CAF50"       # Sýtejšia zelená
GREEN_ACTIVE_BG = "#388E3C"   # Tmavšia zelená pri kliknutí

BLUE_COLOR = "#2196F3"        # Sýtejšia modrá
BLUE_ACTIVE_BG = "#1976D2"    # Tmavšia modrá pri kliknutí
BLUE_TEXT_COLOR = "#FFFFFF"   # Biela pre text na modrých tlačidlách


class App:
    def __init__(self, root: ctk.CTk):
        self.root = root
        self.root.title("Voltoia - Hlavný Nástroj")
        self.root.geometry("850x650")

        # Nastavenie vzhľadu pre customtkinter
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.log_queue = queue.Queue()

        # --- GUI prvky ---
        self.register_pdf_fonts()

        main_frame = ctk.CTkFrame(root, fg_color=BG_COLOR)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- Horný panel s logom a titulkom ---
        header_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        header_frame.pack(fill=tk.X, pady=(0, 15))

        try:
            logo_path = ADRESAR_SKRIPTU / "logo.png"
            original_image = Image.open(logo_path)
            # Použijeme CTkImage pre správne zobrazenie a škálovanie
            self.logo_image = ctk.CTkImage(
                light_image=original_image, size=(100, 90))
            logo_label = ctk.CTkLabel(
                header_frame, image=self.logo_image, text="")
            logo_label.pack(side=tk.LEFT, padx=(0, 15))
        except Exception as e:
            self.log(f"UPOZORNENIE: Nepodarilo sa načítať logo.png: {e}")

        # --- Panely pre nové rozloženie ---
        left_panel = ctk.CTkFrame(
            main_frame, width=200, fg_color="transparent")
        left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10), pady=5)

        right_panel = ctk.CTkFrame(main_frame, fg_color="transparent")
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, pady=5)

        # --- Tlačidlá v ľavom paneli ---
        self.manage_client_button = ctk.CTkButton(
            left_panel,
            text="Sprava klienta",
            command=self.open_manage_client_window,
            fg_color=BUTTON_COLOR,
            text_color=BUTTON_TEXT_COLOR,
            hover_color=BUTTON_ACTIVE_BG,
            font=(
                'Segoe UI',
                13,
                'bold'))
        self.manage_client_button.pack(pady=(5, 5), fill=tk.X)

        self.view_details_button = ctk.CTkButton(
            left_panel,
            text="Detail dňa (DB)",
            command=self.open_details_window,
            fg_color=BUTTON_COLOR,
            text_color=BUTTON_TEXT_COLOR,
            hover_color=BUTTON_ACTIVE_BG,
            font=(
                'Segoe UI',
                13,
                'bold'))
        self.view_details_button.pack(pady=(0, 10), fill=tk.X)

        separator = ctk.CTkFrame(left_panel, height=2, fg_color=LOG_BG_COLOR)
        separator.pack(fill='x', pady=10)

        self.list_active_button = ctk.CTkButton(
            left_panel,
            text="Vypísať aktívnych",
            command=self.list_active_clients,
            fg_color=BLUE_COLOR,
            text_color=BLUE_TEXT_COLOR,
            hover_color=BLUE_ACTIVE_BG)
        self.list_active_button.pack(pady=5, fill=tk.X)
        self.list_inactive_button = ctk.CTkButton(
            left_panel,
            text="Vypísať neaktívnych",
            command=self.list_inactive_clients,
            fg_color=BLUE_COLOR,
            text_color=BLUE_TEXT_COLOR,
            hover_color=BLUE_ACTIVE_BG)
        self.list_inactive_button.pack(pady=5, fill=tk.X)

        self.clear_log_button = ctk.CTkButton(
            left_panel,
            text="Vymazať log",
            command=self.clear_log,
            fg_color=BUTTON_COLOR,
            text_color=BUTTON_TEXT_COLOR,
            hover_color=BUTTON_ACTIVE_BG)
        self.clear_log_button.pack(side=tk.BOTTOM, pady=(10, 5), fill=tk.X)

        # --- Prvky v pravom paneli ---
        self.run_analysis_button = ctk.CTkButton(
            right_panel,
            text="Spustiť hlavnú analýzu",
            command=self.start_analysis_thread,
            fg_color=GREEN_COLOR,
            text_color=BUTTON_TEXT_COLOR,
            hover_color=GREEN_ACTIVE_BG,
            font=(
                'Segoe UI',
                16,
                'bold'),
            height=50)
        self.run_analysis_button.pack(pady=5)

        # Rámček pre tieň okolo logovacieho okna
        log_frame = ctk.CTkFrame(
            right_panel,
            fg_color=BG_COLOR,
            border_width=0)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0), padx=2)

        self.log_area = ctk.CTkTextbox(
            log_frame,
            wrap=tk.WORD,
            state='disabled',
            fg_color=LOG_BG_COLOR,
            text_color=TEXT_COLOR,
            font=(
                'Consolas',
                11),
            border_width=1,
            border_color="#333333")
        self.log_area.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

        self.root.after(100, self.process_log_queue)

    def clear_log(self):
        """Vymaže obsah logovacieho okna."""
        self.log_area.configure(state='normal')
        self.log_area.delete('1.0', tk.END)
        self.log_area.configure(state='disabled')
        self.log("--- Log bol vymazaný ---")

    def register_pdf_fonts(self):
        """
        Skontroluje a zaregistruje písma potrebné pre generovanie PDF.
        Spúšťa sa len raz pri štarte aplikácie.
        """
        font_path = ADRESAR_SKRIPTU / "DejaVuSans.ttf"
        font_bold_path = ADRESAR_SKRIPTU / "DejaVuSans-Bold.ttf"

        if not font_path.exists() or not font_bold_path.exists():
            error_msg = "Súbory DejaVuSans.ttf a DejaVuSans-Bold.ttf neboli nájdené v hlavnom adresári aplikácie."
            self.log(f"❌ KRITICKÁ CHYBA: {error_msg}")
            self.log(
                "   -> Generovanie PDF nebude fungovať správne. Stiahnite si ich a umiestnite vedľa 'main_app.py'.")
            # Upozornenie pre užívateľa, ale aplikácia pobeží ďalej
            messagebox.showwarning(
                "Chýbajúce písmo",
                f"{error_msg}\n\nGenerovanie PDF reportov nemusí fungovať správne.")
        else:
            pdfmetrics.registerFont(TTFont('DejaVuSans', font_path))
            pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', font_bold_path))

    def list_active_clients(self):
        """Vypíše zoznam aktívnych klientov do logovacieho okna."""
        self.log("\n" + "=" * 20 + " ZOZNAM AKTÍVNYCH KLIENTOV " + "=" * 20)
        clients = get_client_list()
        if not clients:
            self.log("Zoznam aktívnych klientov je prázdny.")
        else:
            for i, client in enumerate(clients, 1):
                self.log(f"{i}. {client}")
        self.log("=" * 64)

    def list_inactive_clients(self):
        """Vypíše zoznam neaktívnych klientov do logovacieho okna."""
        self.log("\n" + "=" * 20 + " ZOZNAM NEAKTÍVNYCH KLIENTOV " + "=" * 20)
        clients = get_inactive_client_list()
        if not clients:
            self.log("Zoznam neaktívnych klientov je prázdny.")
        else:
            for i, client in enumerate(clients, 1):
                self.log(f"{i}. {client}")
        self.log("=" * 66)

    def open_manage_client_window(self):
        """Otvorí nové okno s tlačidlami pre správu klientov."""
        self.manage_client_button.configure(state='disabled')

        manage_window = ctk.CTkToplevel(self.root)
        manage_window.title("Sprava klienta")
        manage_window.geometry("300x350")

        frame = ctk.CTkFrame(manage_window, fg_color="transparent")
        frame.pack(expand=True, fill='both', padx=20, pady=20)

        # Wrapper funkcie pre zatvorenie okna pred otvorením ďalšieho
        def wrap_action(action_func):
            def inner():
                manage_window.destroy()
                action_func()
            return inner

        add_button = ctk.CTkButton(
            frame,
            text="Pridať nového klienta",
            command=wrap_action(
                self.open_add_client_window),
            fg_color=GREEN_COLOR,
            text_color=BUTTON_TEXT_COLOR,
            hover_color=GREEN_ACTIVE_BG)
        add_button.pack(pady=5, fill='x')

        edit_button = ctk.CTkButton(
            frame,
            text="Upraviť klienta",
            command=wrap_action(
                self.open_edit_client_window),
            fg_color=BUTTON_COLOR,
            text_color=BUTTON_TEXT_COLOR,
            hover_color=BUTTON_ACTIVE_BG)
        edit_button.pack(pady=5, fill='x')

        delete_button = ctk.CTkButton(
            frame,
            text="Odstrániť klienta",
            command=wrap_action(
                self.open_delete_client_window),
            fg_color="#D32F2F",
            text_color=TEXT_COLOR,
            hover_color="#B71C1C")  # Výrazná červená pre odstránenie
        delete_button.pack(pady=5, fill='x')

        restore_button = ctk.CTkButton(
            frame,
            text="Obnoviť klienta",
            command=wrap_action(
                self.open_restore_client_window),
            fg_color=BLUE_COLOR,
            text_color=BLUE_TEXT_COLOR,
            hover_color=BLUE_ACTIVE_BG)

        restore_button.pack(pady=5, fill='x')

        separator2 = ctk.CTkFrame(frame, height=2, fg_color=LOG_BG_COLOR)
        separator2.pack(fill='x', pady=10)

        permanent_delete_button = ctk.CTkButton(
            frame,
            text="Vymazať neaktívnych (NAVŽDY)",
            command=wrap_action(
                self.open_permanent_delete_client_window),
            fg_color="#8B0000",
            text_color=TEXT_COLOR,
            hover_color="#4B0000")
        permanent_delete_button.pack(pady=5, fill='x')

        open_dir_button = ctk.CTkButton(
            frame,
            text="Otvoriť adresár klienta",
            command=wrap_action(
                self.open_client_directory_window),
            fg_color=BLUE_COLOR,
            text_color=BLUE_TEXT_COLOR,
            hover_color=BLUE_ACTIVE_BG)
        open_dir_button.pack(pady=(15, 5), fill='x')

        open_dir_button = ctk.CTkButton(
            frame,
            text="Prehľad klienta",
            command=wrap_action(
                self.open_client_summary_window),
            fg_color=BUTTON_COLOR,
            text_color=BUTTON_TEXT_COLOR,
            hover_color=BUTTON_ACTIVE_BG)
        open_dir_button.pack(pady=(15, 5), fill='x')

        def on_close():
            self.manage_client_button.configure(state='normal')
            manage_window.destroy()

        manage_window.protocol("WM_DELETE_WINDOW", on_close)
        manage_window.transient(self.root)
        manage_window.grab_set()
        self.root.wait_window(manage_window)
        # Po zatvorení okna sa tlačidlo znovu aktivuje, ak nebolo otvorené
        # ďalšie okno
        if self.manage_client_button.cget('state') == 'disabled':
            self.manage_client_button.configure(state='normal')

    def open_permanent_delete_client_window(self):
        """Otvorí okno pre výber neaktívneho klienta na trvalé vymazanie."""
        delete_window = ctk.CTkToplevel(self.root)
        delete_window.title("Trvalé vymazanie (NAVŽDY)")
        delete_window.geometry("400x400")

        frame = ctk.CTkFrame(delete_window, fg_color="transparent")
        frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

        ctk.CTkLabel(
            frame,
            text="Vyberte neaktívneho klienta na TRVALÉ vymazanie:",
            text_color="#D32F2F",
            font=(
                'Segoe UI',
                12,
                'bold')).pack(
            pady=5)

        client_listbox = Listbox(
            frame,
            selectmode=SINGLE,
            width=50,
            bg=LOG_BG_COLOR,
            fg=TEXT_COLOR,
            relief="flat",
            borderwidth=0,
            highlightthickness=1,
            highlightbackground=BUTTON_COLOR,
            font=(
                'Segoe UI',
                10),
            selectbackground=BUTTON_COLOR,
            selectforeground=BUTTON_TEXT_COLOR)
        client_listbox.pack(pady=5, padx=5, fill=tk.BOTH, expand=True)

        clients = get_inactive_client_list()
        if not clients:
            client_listbox.insert(
                END, "Zoznam neaktívnych klientov je prázdny.")
        else:
            for client in clients:
                client_listbox.insert(END, client)

        def confirm_and_permanently_delete():
            selected_indices = client_listbox.curselection()
            if not selected_indices:  # type: ignore
                messagebox.showwarning(
                    "Žiadny výber",
                    "Prosím, vyberte klienta zo zoznamu.",
                    parent=delete_window)
                return

            client_to_delete = client_listbox.get(selected_indices[0])

            is_confirmed = messagebox.askyesno(
                "KRITICKÉ VAROVANIE",
                f"Naozaj chcete NAVŽDY vymazať klienta '{client_to_delete}' a všetky jeho dáta z databázy?\n\nTúto operáciu nie je možné vrátiť späť!",
                icon='warning',
                parent=delete_window)

            if is_confirmed:
                threading.Thread(
                    target=run_permanently_delete_single_client, args=(
                        client_to_delete, self.log), daemon=True).start()
                delete_window.destroy()

        delete_button = ctk.CTkButton(
            frame,
            text="NAVŽDY vymazať vybraného",
            command=confirm_and_permanently_delete,
            fg_color="#D32F2F",
            text_color=TEXT_COLOR,
            hover_color="#B71C1C")
        delete_button.pack(pady=10, fill='x')

        def on_close():
            delete_window.destroy()

        delete_window.protocol("WM_DELETE_WINDOW", on_close)
        delete_window.transient(self.root)
        delete_window.grab_set()

    def log(self, message):
        self.log_queue.put(message)

    def open_add_client_window(self):
        client_window = ctk.CTkToplevel(self.root)
        client_window.title("Pridať klienta")

        frame = ctk.CTkFrame(client_window, fg_color="transparent")
        frame.pack(expand=True, fill='x', padx=20, pady=20)

        ctk.CTkLabel(frame, text="Mesto:").pack(pady=(0, 2), anchor='w')
        city_entry = ctk.CTkEntry(frame, width=250)
        city_entry.pack(fill='x')

        ctk.CTkLabel(frame, text="Meno:").pack(pady=(10, 2), anchor='w')
        first_name_entry = ctk.CTkEntry(frame, width=250)
        first_name_entry.pack(fill='x')

        ctk.CTkLabel(frame, text="Priezvisko:").pack(pady=(10, 2), anchor='w')
        last_name_entry = ctk.CTkEntry(frame, width=250)
        last_name_entry.pack(fill='x')

        ctk.CTkLabel(
            frame,
            text="Kapacita batérie v kWh:").pack(
            pady=(
                10,
                2),
            anchor='w')
        battery_capacity_entry = ctk.CTkEntry(frame, width=250)
        battery_capacity_entry.pack(fill='x')

        def submit_client():
            # Mesto necháme tak, ako ho zadal užívateľ, geocoding si ho
            # spracuje
            city = city_entry.get().strip()
            first_name = first_name_entry.get().strip().capitalize()
            last_name = last_name_entry.get().strip().capitalize()
            battery_capacity_str = battery_capacity_entry.get().strip().replace(',', '.')

            if not all([city, first_name, last_name, battery_capacity_str]):
                messagebox.showerror(
                    "Chyba",
                    "Všetky polia musia byť vyplnené.",
                    parent=client_window)
                return

            try:
                battery_capacity_kwh = float(battery_capacity_str)
            except ValueError:
                messagebox.showerror(
                    "Chyba",
                    "Kapacita batérie musí byť platné číslo.",
                    parent=client_window)
                return

            # Spustíme pridávanie klienta v novom vlákne, aby GUI nezamrzlo
            threading.Thread(
                target=run_add_client,
                args=(
                    city,
                    first_name,
                    last_name,
                    battery_capacity_kwh,
                    self.log),
                daemon=True).start()
            client_window.destroy()

        submit_button = ctk.CTkButton(
            frame,
            text="Vytvoriť klienta",
            command=submit_client,
            fg_color=GREEN_COLOR,
            text_color=BUTTON_TEXT_COLOR,
            hover_color=GREEN_ACTIVE_BG)
        submit_button.pack(pady=20, fill='x')

        def on_close():
            client_window.destroy()

        client_window.protocol("WM_DELETE_WINDOW", on_close)
        client_window.transient(self.root)
        client_window.grab_set()

    def open_edit_client_window(self):
        """Otvorí okno pre výber a následnú úpravu klienta."""

        # --- Krok 1: Okno na výber klienta ---
        select_window = ctk.CTkToplevel(self.root)
        select_window.title("Vybrať klienta na úpravu")
        select_window.geometry("400x400")

        frame = ctk.CTkFrame(select_window, fg_color="transparent")
        frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

        ctk.CTkLabel(
            frame,
            text="Vyberte klienta, ktorého údaje chcete upraviť:").pack(
            pady=5)

        client_listbox = Listbox(
            frame,
            selectmode=SINGLE,
            width=50,
            bg=LOG_BG_COLOR,
            fg=TEXT_COLOR,
            relief="flat",
            borderwidth=0,
            highlightthickness=1,
            highlightbackground=BUTTON_COLOR,
            font=(
                'Segoe UI',
                10),
            selectbackground=BUTTON_COLOR,
            selectforeground=BUTTON_TEXT_COLOR)
        client_listbox.pack(pady=5, padx=5, fill=tk.BOTH, expand=True)

        clients = get_client_list()
        if not clients:
            client_listbox.insert(END, "Zoznam aktívnych klientov je prázdny.")
        else:
            for client in clients:
                client_listbox.insert(END, client)

        def on_select_and_open_form():
            selected_indices = client_listbox.curselection()
            if not selected_indices:  # type: ignore
                messagebox.showwarning(
                    "Žiadny výber",
                    "Prosím, vyberte klienta zo zoznamu.",
                    parent=select_window)
                return

            client_identifier = client_listbox.get(selected_indices[0])
            select_window.destroy()
            self.open_client_edit_form(client_identifier)

        select_button = ctk.CTkButton(
            frame,
            text="Upraviť vybraného klienta",
            command=on_select_and_open_form,
            fg_color=GREEN_COLOR,
            text_color=BUTTON_TEXT_COLOR,
            hover_color=GREEN_ACTIVE_BG)
        select_button.pack(pady=10, fill='x')

        def on_close():
            select_window.destroy()

        select_window.protocol("WM_DELETE_WINDOW", on_close)
        select_window.transient(self.root)
        select_window.grab_set()  # grab_set() je dôležité, aby bolo okno modálne

    def open_client_edit_form(self, client_identifier):
        """Otvorí editačný formulár s predvyplnenými údajmi klienta."""
        client_data = get_client_details(client_identifier)
        if not client_data:
            messagebox.showerror(
                "Chyba", f"Nepodarilo sa načítať údaje pre klienta '{client_identifier}'.")
            return

        edit_window = ctk.CTkToplevel(self.root)
        edit_window.title(f"Úprava klienta")

        frame = ctk.CTkFrame(edit_window, fg_color="transparent")
        frame.pack(expand=True, fill='x', padx=20, pady=20)

        ctk.CTkLabel(
            frame,
            text=f"Úprava údajov pre: {client_identifier}",
            font=(
                'Segoe UI',
                12,
                'bold')).pack(
            pady=(
                0,
                15),
            anchor='w')

        # Meno
        ctk.CTkLabel(frame, text="Meno:").pack(anchor='w')
        first_name_entry = ctk.CTkEntry(frame, width=300)
        first_name_entry.insert(0, client_data.get('meno', ''))
        first_name_entry.pack(fill='x', pady=(0, 10))

        # Priezvisko
        ctk.CTkLabel(frame, text="Priezvisko:").pack(anchor='w')
        last_name_entry = ctk.CTkEntry(frame, width=300)
        last_name_entry.insert(0, client_data.get('priezvisko', ''))
        last_name_entry.pack(fill='x', pady=(0, 10))

        # Mesto
        ctk.CTkLabel(frame, text="Mesto:").pack(anchor='w')
        city_entry = ctk.CTkEntry(frame, width=300)
        city_entry.insert(0, client_data.get('mesto', ''))
        city_entry.pack(fill='x', pady=(0, 10))

        # Kapacita batérie
        ctk.CTkLabel(frame, text="Kapacita batérie v kWh:").pack(anchor='w')
        battery_capacity_entry = ctk.CTkEntry(frame, width=300)
        battery_capacity_entry.insert(
            0, str(
                client_data.get(
                    'kapacita batérie v kWh', '')))
        battery_capacity_entry.pack(fill='x', pady=(0, 10))

        def submit_changes():
            try:
                battery_capacity_val = float(
                    battery_capacity_entry.get().strip().replace(
                        ',', '.'))
                new_data = {
                    'meno': first_name_entry.get().strip(),
                    'priezvisko': last_name_entry.get().strip().capitalize(),
                    'mesto': normalizuj_mesto(city_entry.get().strip()),
                    'kapacita batérie v kWh': battery_capacity_val
                }
            except ValueError:
                messagebox.showerror(
                    "Chyba",
                    "Kapacita batérie musí byť platné číslo.",
                    parent=edit_window)
                return
            # Spustíme úpravu priamo, nie v novom vlákne, aby sme predišli
            # problémom s GUI
            run_edit_client(client_identifier, new_data, self.log)
            edit_window.destroy()

        save_button = ctk.CTkButton(frame, text="Uložiť zmeny", command=submit_changes,  # type: ignore
                                    fg_color=GREEN_COLOR, text_color=BUTTON_TEXT_COLOR, hover_color=GREEN_ACTIVE_BG)
        save_button.pack(pady=(20, 0), fill='x')  # Tento riadok chýbal

        def on_edit_close():
            edit_window.destroy()

        edit_window.protocol("WM_DELETE_WINDOW", on_edit_close)

    def open_delete_client_window(self):

        # Používame CTkToplevel pre konzistentný vzhľad
        delete_window = ctk.CTkToplevel(self.root)
        delete_window.title("Odstrániť klienta")
        delete_window.geometry("400x400")

        frame = ctk.CTkFrame(delete_window, fg_color="transparent")
        frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

        ctk.CTkLabel(
            frame,
            text="Vyberte klienta na odstránenie:").pack(
            pady=5)

        client_listbox = Listbox(
            frame,
            selectmode=SINGLE,
            width=50,
            bg=LOG_BG_COLOR,
            fg=TEXT_COLOR,
            relief="flat",
            borderwidth=0,
            highlightthickness=1,
            highlightbackground=BUTTON_COLOR,
            font=(
                'Segoe UI',
                10),
            selectbackground=BUTTON_COLOR,
            selectforeground=BUTTON_TEXT_COLOR)
        client_listbox.pack(pady=5, padx=5, fill=tk.BOTH, expand=True)

        clients = get_client_list()
        if not clients:
            client_listbox.insert(END, "Zoznam aktívnych klientov je prázdny.")
        else:
            for client in clients:
                client_listbox.insert(END, client)

        def confirm_and_delete():
            selected_indices = client_listbox.curselection()
            if not selected_indices:  # type: ignore
                messagebox.showwarning(
                    "Žiadny výber",
                    "Prosím, vyberte klienta zo zoznamu.",
                    parent=delete_window)
                return

            client_to_delete = client_listbox.get(selected_indices[0])

            is_confirmed = messagebox.askyesno(
                "Potvrdenie odstránenia",
                f"Naozaj si prajete deaktivovať klienta '{client_to_delete}'?\n\nKlient bude presunutý do zoznamu neaktívnych.",
                icon='warning',
                parent=delete_window)

            if is_confirmed:
                threading.Thread(
                    target=run_delete_client,
                    args=(
                        client_to_delete,
                        self.log),
                    daemon=True).start()
                delete_window.destroy()

        delete_button = ctk.CTkButton(
            frame,
            text="Odstrániť vybraného klienta",
            command=confirm_and_delete,
            fg_color="#D32F2F",
            text_color=TEXT_COLOR,
            hover_color="#B71C1C")
        delete_button.pack(pady=10, fill='x')

        def on_close():
            """Zabezpečí, že sa tlačidlo vždy re-aktivuje."""
            delete_window.destroy()

        delete_window.protocol("WM_DELETE_WINDOW", on_close)
        delete_window.transient(self.root)
        delete_window.grab_set()

    def open_restore_client_window(self):

        restore_window = ctk.CTkToplevel(self.root)
        restore_window.title("Obnoviť klienta")
        restore_window.geometry("400x400")

        frame = ctk.CTkFrame(restore_window, fg_color="transparent")
        frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

        ctk.CTkLabel(
            frame,
            text="Vyberte neaktívneho klienta na obnovenie:").pack(
            pady=5)

        client_listbox = Listbox(
            frame,
            selectmode=SINGLE,
            width=50,
            bg=LOG_BG_COLOR,
            fg=TEXT_COLOR,
            relief="flat",
            borderwidth=0,
            highlightthickness=1,
            highlightbackground=BUTTON_COLOR,
            font=(
                'Segoe UI',
                10),
            selectbackground=BUTTON_COLOR,
            selectforeground=BUTTON_TEXT_COLOR)
        client_listbox.pack(pady=5, padx=5, fill=tk.BOTH, expand=True)

        clients = get_inactive_client_list()
        if not clients:
            client_listbox.insert(
                END, "Zoznam neaktívnych klientov je prázdny.")
        else:
            for client in clients:
                client_listbox.insert(END, client)

        def confirm_and_restore():
            selected_indices = client_listbox.curselection()
            if not selected_indices:  # type: ignore
                messagebox.showwarning(
                    "Žiadny výber",
                    "Prosím, vyberte klienta zo zoznamu.",
                    parent=restore_window)
                return

            client_to_restore = client_listbox.get(selected_indices[0])

            is_confirmed = messagebox.askyesno(
                "Potvrdenie obnovenia",
                f"Naozaj si prajete obnoviť klienta '{client_to_restore}'?",
                parent=restore_window
            )

            if is_confirmed:
                threading.Thread(
                    target=run_restore_client,
                    args=(
                        client_to_restore,
                        self.log),
                    daemon=True).start()
                restore_window.destroy()

        restore_button = ctk.CTkButton(
            frame,
            text="Obnoviť vybraného klienta",
            command=confirm_and_restore,
            fg_color=GREEN_COLOR,
            text_color=BUTTON_TEXT_COLOR,
            hover_color=GREEN_ACTIVE_BG)
        restore_button.pack(pady=10, fill='x')

        def on_close():
            restore_window.destroy()

        restore_window.protocol("WM_DELETE_WINDOW", on_close)
        restore_window.transient(self.root)
        restore_window.grab_set()

    def open_client_directory_window(self):
        """Otvorí okno pre výber klienta, ktorého adresár sa má otvoriť."""
        select_window = ctk.CTkToplevel(self.root)
        select_window.title("Otvoriť adresár klienta")
        select_window.geometry("400x400")

        frame = ctk.CTkFrame(select_window, fg_color="transparent")
        frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

        ctk.CTkLabel(
            frame,
            text="Vyberte klienta, ktorého adresár chcete otvoriť:").pack(
            pady=5)

        client_listbox = Listbox(
            frame,
            selectmode=SINGLE,
            width=50,
            bg=LOG_BG_COLOR,
            fg=TEXT_COLOR,
            relief="flat",
            borderwidth=0,
            highlightthickness=1,
            highlightbackground=BUTTON_COLOR,
            font=(
                'Segoe UI',
                10),
            selectbackground=BUTTON_COLOR,
            selectforeground=BUTTON_TEXT_COLOR)
        client_listbox.pack(pady=5, padx=5, fill=tk.BOTH, expand=True)

        clients = get_client_list()
        if not clients:
            client_listbox.insert(END, "Zoznam aktívnych klientov je prázdny.")
        else:
            for client in clients:
                client_listbox.insert(END, client)

        def on_select_and_open():
            selected_indices = client_listbox.curselection()
            if not selected_indices:  # type: ignore
                messagebox.showwarning(
                    "Žiadny výber",
                    "Prosím, vyberte klienta zo zoznamu.",
                    parent=select_window)
                return

            client_identifier = client_listbox.get(selected_indices[0])
            client_data = get_client_details(client_identifier)
            if client_data and client_data.get('mesto'):
                client_dir_path = ADRESAR_SKRIPTU / \
                    normalizuj_mesto(client_data['mesto']) / client_identifier
                open_directory(client_dir_path, self.log)
            else:
                self.log(
                    f"❌ CHYBA: Nepodarilo sa získať údaje (mesto) pre klienta '{client_identifier}'.")
            select_window.destroy()

        select_button = ctk.CTkButton(
            frame,
            text="Otvoriť adresár",
            command=on_select_and_open,
            fg_color=BLUE_COLOR,
            text_color=BLUE_TEXT_COLOR,
            hover_color=BLUE_ACTIVE_BG)
        select_button.pack(pady=10, fill='x')

    def open_client_summary_window(self):
        """Otvorí okno pre výber klienta, pre ktorého sa zobrazí súhrn."""
        select_window = ctk.CTkToplevel(self.root)
        select_window.title("Zobraziť prehľad klienta")
        select_window.geometry("400x400")

        frame = ctk.CTkFrame(select_window, fg_color="transparent")
        frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

        ctk.CTkLabel(
            frame,
            text="Vyberte klienta, ktorého prehľad chcete zobraziť:").pack(
            pady=5)

        client_listbox = Listbox(
            frame,
            selectmode=SINGLE,
            width=50,
            bg=LOG_BG_COLOR,
            fg=TEXT_COLOR,
            relief="flat",
            borderwidth=0,
            highlightthickness=1,
            highlightbackground=BUTTON_COLOR,
            font=(
                'Segoe UI',
                10),
            selectbackground=BUTTON_COLOR,
            selectforeground=BUTTON_TEXT_COLOR)
        client_listbox.pack(pady=5, padx=5, fill=tk.BOTH, expand=True)

        clients = get_client_list()
        if not clients:
            client_listbox.insert(END, "Zoznam aktívnych klientov je prázdny.")
        else:
            for client in clients:
                client_listbox.insert(END, client)

        def on_select_and_show():
            selected_indices = client_listbox.curselection()
            if not selected_indices:  # type: ignore
                messagebox.showwarning(
                    "Žiadny výber",
                    "Prosím, vyberte klienta zo zoznamu.",
                    parent=select_window)
                return

            client_identifier = client_listbox.get(selected_indices[0])
            client_data = get_client_details(client_identifier)
            if client_data and client_data.get('mesto'):
                client_dir_path = ADRESAR_SKRIPTU / \
                    normalizuj_mesto(client_data['mesto']) / client_identifier
                summary_data, df_full = calculate_client_summary(
                    client_dir_path, self.log)
                select_window.destroy()
                self.display_summary_window(
                    client_identifier, summary_data, df_full, client_data)
            else:
                self.log(
                    f"❌ CHYBA: Nepodarilo sa získať údaje (mesto) pre klienta '{client_identifier}'.")
                select_window.destroy()

        select_button = ctk.CTkButton(
            frame,
            text="Zobraziť prehľad",
            command=on_select_and_show,
            fg_color=BUTTON_COLOR,
            text_color=BUTTON_TEXT_COLOR,
            hover_color=BUTTON_ACTIVE_BG)
        select_button.pack(pady=10, fill='x')

    def display_summary_window(
            self,
            client_identifier,
            summary_data,
            df_full,
            client_data):
        """Zobrazí okno so súhrnnými dátami klienta."""
        summary_window = ctk.CTkToplevel(self.root)
        summary_window.title(
            f"Mesacny prehlad pre klienta : {client_identifier}")
        summary_window.minsize(width=550, height=400)

        frame = ctk.CTkFrame(summary_window, fg_color="transparent")
        frame.pack(expand=True, fill='both', padx=20, pady=20)

        if "error" in summary_data:
            ctk.CTkLabel(
                frame,
                text=f"Chyba: {
                    summary_data['error']}",
                text_color="red",
                font=(
                    'Segoe UI',
                    14,
                    'bold')).pack(
                pady=10)
            return

        # Pridanie kapacity batérie
        battery_capacity = client_data.get('kapacita batérie v kWh', 'N/A')
        row_frame_capacity = ctk.CTkFrame(frame, fg_color="transparent")
        row_frame_capacity.pack(fill='x', pady=2)
        ctk.CTkLabel(
            row_frame_capacity,
            text="Kapacita batérie (kWh):",
            font=(
                'Segoe UI',
                12,
                'bold'),
            anchor='w').pack(
            side='left')
        ctk.CTkLabel(
            row_frame_capacity,
            text=str(battery_capacity),
            font=(
                'Segoe UI',
                12),
            anchor='e').pack(
            side='right',
            padx=(
                10,
                 0))

        # Define keys to hide from the summary window
        hidden_keys = {
            "Rozdiel nakup / predaj",
            "Potenciálny brutto zisk (sum cien predaja)",
            "Počet ziskových dní",
            "Počet stratových dní",
        }

        # Zobrazenie štandardných metrík
        for key, value in summary_data.items():
            if key not in hidden_keys and key != "Štatistika FVE":
                row_frame = ctk.CTkFrame(frame, fg_color="transparent")
                row_frame.pack(fill='x', pady=2)
                ctk.CTkLabel(
                    row_frame,
                    text=f"{key}:",
                    font=(
                        'Segoe UI',
                        12,
                        'bold'),
                    anchor='w').pack(
                    side='left')
                ctk.CTkLabel(
                    row_frame, text=value, font=(
                        'Segoe UI', 12), anchor='e').pack(
                    side='right', padx=(
                        10, 0))

        # Zobrazenie štatistiky FVE
        fve_stats = summary_data.get("Štatistika FVE", {})
        if fve_stats:
            ctk.CTkLabel(
                frame,
                text="Štatistika FVE:",
                font=(
                    'Segoe UI',
                    12,
                    'bold'),
                anchor='w').pack(
                fill='x',
                pady=(
                    15,
                    5))
            fve_text = "\n".join(
                [f"  - {k.capitalize() if isinstance(k, str) else k}: {v} dní" for k, v in fve_stats.items()])
            ctk.CTkLabel(
                frame,
                text=fve_text,
                font=(
                    'Segoe UI',
                    12),
                justify='left',
                anchor='w').pack(
                fill='x')

        # Tlačidlo na generovanie PDF
        pdf_button = ctk.CTkButton(
            frame,
            text="Generovať PDF report",
            command=lambda: self.start_pdf_generation(
                client_identifier,
                summary_data,
                df_full,
                client_data),
            fg_color=BLUE_COLOR,
            text_color=BLUE_TEXT_COLOR,
            hover_color=BLUE_ACTIVE_BG)
        pdf_button.pack(pady=(20, 5), fill='x')

    def start_pdf_generation(
            self,
            client_identifier,
            summary_data,
            df_full,
            client_data):
        """Spustí generovanie PDF v samostatnom vlákne."""
        self.log(
            f"Spúšťam generovanie PDF reportu pre klienta '{client_identifier}'...")
        thread = threading.Thread(
            target=self.generate_pdf_report,
            args=(
                client_identifier,
                summary_data,
                df_full,
                client_data),
            daemon=True)
        thread.start()

    def generate_pdf_report(
            self,
            client_identifier,
            summary_data,
            df_full,
            client_data):
        """Vygeneruje PDF report s dátami a grafom."""
        try:
            # --- Príprava cesty pre uloženie PDF ---
            if not client_data or not client_data.get('mesto'):
                self.log(
                    f"❌ CHYBA: Nepodarilo sa získať adresár pre uloženie PDF pre klienta '{client_identifier}'.")
                return

            client_dir_path = ADRESAR_SKRIPTU / \
                normalizuj_mesto(client_data['mesto']) / client_identifier
            pdf_filename = f"Report_{client_identifier}_{
                datetime.date.today().strftime('%Y-%m')}.pdf"
            pdf_path = client_dir_path / pdf_filename

            # --- Generovanie grafu ---
            # Kľúčová oprava: Povieme Matplotlibu, aby použil neinteraktívny
            # backend, čo je bezpečné pre vlákna.
            import matplotlib
            matplotlib.use('Agg')

            # Používame nový stĺpec pre celkový zisk
            profit_col = 'celkový zisk (€)'
            df_full['dátum_dt'] = pd.to_datetime(
                df_full['dátum'], format='%d.%m.%Y')  # type: ignore

            fig, ax = plt.subplots(figsize=(10, 5))
            colors = ['#4CAF50' if x >
                      0 else '#F44336' for x in df_full[profit_col]]
            ax.bar(df_full['dátum_dt'], df_full[profit_col], color=colors)

            ax.set_title(
                'Denný celkový zisk',
                fontsize=14)  # Upravený názov grafu
            ax.set_ylabel('Zisk v €', fontsize=10)
            ax.grid(axis='y', linestyle='--', alpha=0.7)
            ax.axhline(0, color='black', linewidth=0.8)
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%d.%m'))
            plt.xticks(rotation=45, ha="right")
            plt.tight_layout()

            graph_path = client_dir_path / "temp_graph.png"
            plt.savefig(graph_path, dpi=150)
            plt.close(fig)

            # --- Generovanie PDF ---
            c = canvas.Canvas(str(pdf_path), pagesize=A4)
            width, height = A4

            # Hlavička
            c.setFont("DejaVuSans", 14)
            c.drawString(
                30,
                height - 50,
                f"Mesačný report pre klienta: {client_identifier}")
            c.setFont("DejaVuSans", 12)
            c.drawString(
                30,
                height - 65,
                f"Dátum generovania: {
                    datetime.date.today().strftime('%d.%m.%Y')}")

            # Pridanie loga (ak existuje)
            logo_path = ADRESAR_SKRIPTU / "logo_transparent.png"
            if logo_path.exists():
                try:
                    # Umiestnenie loga vpravo hore (cca 80x70)
                    logo_width = 80
                    logo_height = 70
                    c.drawImage(
                        ImageReader(
                            str(logo_path)),
                        width - 30 - logo_width,
                        height - 85,
                        width=logo_width,
                        height=logo_height,
                        preserveAspectRatio=True,
                        mask='auto')
                except Exception as e:
                    self.log(f"⚠️ Nepodarilo sa pridať logo do PDF: {e}")

            c.line(30, height - 75, width - 30, height - 75)

            # Textové dáta
            styles = getSampleStyleSheet()
            style_normal = ParagraphStyle(
                'BodyTextCustom',
                parent=styles['BodyText'],
                fontName='DejaVuSans',
                fontSize=12,
                leading=15)

            y_pos = height - 100

            # Pridanie kapacity batérie
            battery_capacity = client_data.get('kapacita batérie v kWh', 'N/A')
            p = Paragraph(
                f"<font name='DejaVuSans-Bold'>Kapacita batérie (kWh):</font> {battery_capacity}",
                style_normal)
            p_width, p_height = p.wrapOn(c, width - 60, 20)
            p.drawOn(c, 30, y_pos)
            y_pos -= p_height + 5

            # Rovnaký filter ako v GUI
            hidden_keys = {
                "Rozdiel nakup / predaj",
                "Potenciálny brutto zisk (sum cien predaja)",
                "Počet ziskových dní",
                "Počet stratových dní",
            }

            for key, value in summary_data.items():
                if key not in hidden_keys and key != "Štatistika FVE":
                    p = Paragraph(
                        f"<font name='DejaVuSans-Bold'>{key}:</font> {value}",
                        style_normal)
                    p_width, p_height = p.wrapOn(c, width - 60, 20)
                    p.drawOn(c, 30, y_pos)
                    y_pos -= p_height + 5

            # Graf
            c.drawImage(
                ImageReader(graph_path),
                30,
                y_pos - 280,
                width=width - 60,
                height=250,
                preserveAspectRatio=True)
            c.save()
            os.remove(graph_path)  # Zmažeme dočasný súbor s grafom

            self.log(
                f"✅ PDF report '{pdf_filename}' bol úspešne vygenerovaný a uložený.")
            messagebox.showinfo(
                "Úspech",
                f"PDF report bol úspešne vygenerovaný a uložený v adresári klienta:\n\n{pdf_path}")

        except Exception as e:
            self.log(f"❌ CHYBA pri generovaní PDF reportu: {e}")
            messagebox.showerror(
                "Chyba", f"Nepodarilo sa vygenerovať PDF report:\n{e}")

    def start_analysis_thread(self):
        self.run_analysis_button.configure(state='disabled')
        self.log("Spúšťam analýzu v samostatnom vlákne...")
        analysis_thread = threading.Thread(
            target=self.run_analysis_and_reenable_button, daemon=True)
        analysis_thread.start()

    def run_analysis_and_reenable_button(self):
        try:
            dnes = datetime.date.today()
            zajtra = dnes + datetime.timedelta(days=1)

            teraz = datetime.datetime.now()
            target_dates = [zajtra]

            # Kontrola dostupnosti dát pre zajtrajšok (iba ak je pred 15:00)
            if teraz.hour < 15:
                self.log("Kontrolujem dostupnosť dát pre zajtrajšok na OKTE...")
                def dummy_log(msg): return None
                df_test = ziskaj_okte_data(zajtra, dummy_log)

                # Ak dáta nie sú (None) alebo sú prázdne (len hlavičky)
                if df_test is None or df_test.empty:
                    self.log("⚠️ Dáta pre zajtrajšok ešte nie sú k dispozícii.")

                    def ask_user():
                        return messagebox.askyesno(
                            "Dáta nie sú dostupné",
                            "Dáta pre zajtrajšok ešte nie sú na OKTE uverejnené (zvyčajne po 15:00).\n\n"
                            "Chcete namiesto toho spustiť analýzu pre DNEŠNÝ DEŇ?",
                            parent=self.root)

                    response_event = threading.Event()
                    user_choice = {"analyze_today": False}

                    def handle_choice():
                        user_choice["analyze_today"] = ask_user()
                        response_event.set()

                    self.root.after(0, handle_choice)
                    response_event.wait()

                    if user_choice["analyze_today"]:
                        self.log("👉 Zvolená analýza pre DNEŠOK.")
                        target_dates = [dnes]
                    else:
                        self.log("❌ Analýza zrušená užívateľom.")
                        return

            run_main_analysis(self.log, target_dates=target_dates)
        except Exception as e:
            self.log(f"❌ KRITICKÁ CHYBA VO VLÁKNE ANALÝZY: {e}")
        finally:
            # Musíme poslať zmenu stavu tlačidla cez frontu, lebo meníme GUI z
            # iného vlákna
            self.log_queue.put("REENABLE_BUTTON")

    def process_log_queue(self):
        while not self.log_queue.empty():
            try:
                message = self.log_queue.get_nowait()
                if message == "REENABLE_BUTTON":
                    self.run_analysis_button.configure(state='normal')
                else:
                    self.log_area.configure(state='normal')
                    self.log_area.insert(tk.END, message + '\n')
                    self.log_area.configure(state='disabled')
                    self.log_area.see(tk.END)
            except queue.Empty:
                pass
        self.root.after(100, self.process_log_queue)

    def open_details_window(self):
        """Otvorí okno pre detailné zobrazenie dňa z databázy."""
        details_window = ctk.CTkToplevel(self.root)
        details_window.title("Detail dňa (Databáza)")
        details_window.geometry("900x700")
        details_window.attributes('-topmost', True)

        # Ovládací panel
        control_frame = ctk.CTkFrame(details_window)
        control_frame.pack(fill='x', padx=10, pady=10)

        ctk.CTkLabel(control_frame, text="Klient:").pack(side='left', padx=5)
        self.client_combo = ctk.CTkComboBox(
            control_frame,
            values=get_client_list(),
            width=250,
            command=self.update_dates_combobox)
        self.client_combo.set("")  # Empty by default
        self.client_combo.pack(side='left', padx=5)

        ctk.CTkLabel(control_frame, text="Dátum:").pack(side='left', padx=5)
        self.date_combo = ctk.CTkComboBox(control_frame, values=[], width=150)
        self.date_combo.set("")  # Empty by default
        self.date_combo.pack(side='left', padx=5)

        show_btn = ctk.CTkButton(
            control_frame,
            text="Zobraziť",
            command=self.show_daily_data,
            fg_color=GREEN_COLOR,
            hover_color=GREEN_ACTIVE_BG)
        show_btn.pack(side='left', padx=15)

        # Sumárny panel
        self.summary_frame = ctk.CTkFrame(details_window)
        self.summary_frame.pack(fill='x', padx=10, pady=5)
        self.summary_label = ctk.CTkLabel(
            self.summary_frame,
            text="Vyberte klienta a dátum pre zobrazenie detailov.",
            font=(
                'Segoe UI',
                14))
        self.summary_label.pack(pady=5)

        # Tabuľka (Treeview)
        table_frame = ctk.CTkFrame(details_window)
        table_frame.pack(fill='both', expand=True, padx=10, pady=10)

        columns = ("cas", "cena", "fve", "akcia")
        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show='headings',
            height=20)

        self.tree.heading("cas", text="Čas")
        self.tree.heading("cena", text="Cena (€/MWh)")
        self.tree.heading("fve", text="FVE (kWh)")
        self.tree.heading("akcia", text="Akcia")

        self.tree.column("cas", width=100, anchor='center')
        self.tree.column("cena", width=100, anchor='center')
        self.tree.column("fve", width=100, anchor='center')
        self.tree.column("akcia", width=150, anchor='center')

        # Scrollbar
        scrollbar = ttk.Scrollbar(
            table_frame,
            orient="vertical",
            command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side='right', fill='y')
        self.tree.pack(side='left', fill='both', expand=True)

        # Definícia farieb (tagov)
        self.tree.tag_configure('nakup', background='#C8E6C9')  # Zelená
        self.tree.tag_configure('predaj', background='#BBDEFB')  # Modrá
        self.tree.tag_configure('nabijanie', background='#FFE0B2')  # Oranžová
        # Žltá (pre záporné ceny)
        self.tree.tag_configure('zaporna', background='#FFF9C4')

    def update_dates_combobox(self, choice):
        if not choice:
            return
        client_id = choice.split('_')[0]
        dates = database_manager.get_available_dates(client_id)
        self.date_combo.configure(values=dates)
        if dates:
            self.date_combo.set(dates[0])
        else:
            self.date_combo.set("")

    def show_daily_data(self):
        client_str = self.client_combo.get()
        date_str = self.date_combo.get()

        # Check if client and date are filled
        if not client_str or not date_str:
            messagebox.showwarning(
                "Chýbajúce údaje",
                "Vyberte klienta a dátum.")
            return

        # Check if client is valid (exists in the client list)
        client_list = get_client_list()
        if client_str not in client_list:
            messagebox.showwarning(
                "Chýbajúce údaje",
                "Prosím, vyberte klienta zo zoznamu.")
            return

        client_id = client_str.split('_')[0]
        data = database_manager.get_daily_detail(client_id, date_str)

        if not data:
            messagebox.showerror("Chyba", "Dáta pre tento deň neboli nájdené.")
            return

        # Vyčistenie tabuľky
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Vyčistenie sumárneho panela
        for widget in self.summary_frame.winfo_children():
            widget.destroy()

        # Hlavný súhrn
        profit = data.get('total_profit', 0)
        fve_text = data.get('fve_prediction_text', 'N/A')

        header_label = ctk.CTkLabel(
            self.summary_frame, text=f"📅 Dátum: {date_str}  |  💰 Celkový zisk: {
                profit:.2f} €  |  ☀️ FVE: {fve_text}", font=(
                'Segoe UI', 14, 'bold'))
        header_label.pack(pady=(5, 10))

        # Detailné bloky
        blocks_frame = ctk.CTkFrame(self.summary_frame, fg_color="transparent")
        blocks_frame.pack(fill='x', padx=20)

        # Nadpis
        title_label = ctk.CTkLabel(blocks_frame, text="📊 BLOKY:",
                                   font=('Segoe UI', 13, 'bold'), anchor='w')
        title_label.pack(fill='x', pady=(0, 5))

        # Blok 1 - Nákup (zelený)
        block1 = data.get('block_1_buy', 'N/A')
        price1 = data.get('price_buy_1', 0)
        if block1 and block1 != 'N/A':
            label1 = ctk.CTkLabel(
                blocks_frame,
                text=f"🟢 Nákup 1:      {block1}  ({
                    price1:.2f} €/MWh)",
                font=(
                    'Consolas',
                    12),
                anchor='w',
                text_color="#4CAF50")
            label1.pack(fill='x', pady=2)

        # Blok 2 - Predaj (modrý)
        block2 = data.get('block_2_sell', 'N/A')
        price2 = data.get('price_sell_1', 0)
        if block2 and block2 != 'N/A':
            label2 = ctk.CTkLabel(
                blocks_frame,
                text=f"🔵 Predaj 1:     {block2}  ({
                    price2:.2f} €/MWh)",
                font=(
                    'Consolas',
                    12),
                anchor='w',
                text_color="#2196F3")
            label2.pack(fill='x', pady=2)

        # Blok 3 - FVE nabíjanie (oranžový)
        block3 = data.get('block_3_buy', 'N/A')
        if block3 and block3 != 'N/A':
            label3 = ctk.CTkLabel(
                blocks_frame, text=f"🟠 FVE nab.:     {block3}", font=(
                    'Consolas', 12), anchor='w', text_color="#FF9800")
            label3.pack(fill='x', pady=2)

        # Blok 4 - Predaj 2 (modrý)
        block4 = data.get('block_4_sell', 'N/A')
        price4 = data.get('price_sell_2', 0)
        if block4 and block4 != 'N/A':
            label4 = ctk.CTkLabel(
                blocks_frame,
                text=f"🔵 Predaj 2:     {block4}  ({
                    price4:.2f} €/MWh)",
                font=(
                    'Consolas',
                    12),
                anchor='w',
                text_color="#2196F3")
            label4.pack(fill='x', pady=2)

        # Záporné ceny
        zero_price = data.get('zero_price_intervals', '')
        if zero_price:
            label_zero = ctk.CTkLabel(
                blocks_frame,
                text=f"⚠️  Záporné/0€:  {zero_price}",
                font=(
                    'Consolas',
                    11),
                anchor='w',
                text_color='#FFA500')
            label_zero.pack(fill='x', pady=(8, 2))

        # Naplnenie tabuľky
        hourly_json = data.get('hourly_data_json')
        if hourly_json:
            try:
                hourly_data = json.loads(hourly_json)
                for item in hourly_data:
                    cas = item.get('cas', '')
                    cena = item.get('cena', 0)
                    fve = item.get('fve', 0)
                    akcia_raw = item.get('akcia', 'nic')

                    akcia_text = ""
                    tag = ""

                    if "nakup" in akcia_raw:
                        akcia_text = "Nákup"
                        tag = "nakup"
                    elif "predaj" in akcia_raw:
                        akcia_text = "Predaj"
                        tag = "predaj"
                    elif "nabijanie" in akcia_raw:
                        akcia_text = "Nabíjanie FVE"
                        tag = "nabijanie"

                    if cena <= 0:
                        if not tag:
                            tag = "zaporna"

                    values = (cas, f"{cena:.2f}", f"{fve:.2f}", akcia_text)
                    self.tree.insert('', 'end', values=values, tags=(tag,))
            except json.JSONDecodeError:
                messagebox.showerror(
                    "Chyba", "Chyba pri čítaní detailných dát (JSON).")


if __name__ == "__main__":
    root = ctk.CTk()
    app = App(root)
    root.mainloop()  # Test
